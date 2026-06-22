from datetime import date, timedelta, datetime
from decimal import Decimal
from io import BytesIO
from typing import Any
import re
import unicodedata

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session, joinedload

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from app.core.deps import get_db
from app.models.financeiro_pagar import FinanceiroPagar
from app.models.financeiro_receber import FinanceiroReceber


router = APIRouter(prefix="/api/fluxo-caixa", tags=["Fluxo de Caixa"])


FORMAS_PAGAMENTO_PADRAO = [
    ("DINHEIRO", "Dinheiro"),
    ("PIX", "PIX"),
    ("CONTA_BANCARIA", "Conta bancária"),
    ("CARTAO_CREDITO", "Cartão de crédito"),
    ("CARTAO_DEBITO", "Cartão de débito"),
    ("BOLETO", "Boleto"),
    ("OUTRO", "Outro"),
]

MAPA_FORMAS_PAGAMENTO = {
    "DINHEIRO": "DINHEIRO",
    "PIX": "PIX",
    "CONTA": "CONTA_BANCARIA",
    "CONTA BANCARIA": "CONTA_BANCARIA",
    "CONTABANCARIA": "CONTA_BANCARIA",
    "CONTA_BANCARIA": "CONTA_BANCARIA",
    "BANCO": "CONTA_BANCARIA",
    "BANCARIA": "CONTA_BANCARIA",
    "TRANSFERENCIA": "CONTA_BANCARIA",
    "TRANSFERENCIA BANCARIA": "CONTA_BANCARIA",
    "TED": "CONTA_BANCARIA",
    "DOC": "CONTA_BANCARIA",
    "CARTAO": "CARTAO_CREDITO",
    "CARTAO CREDITO": "CARTAO_CREDITO",
    "CARTAOCREDITO": "CARTAO_CREDITO",
    "CARTAO_CREDITO": "CARTAO_CREDITO",
    "CREDITO": "CARTAO_CREDITO",
    "CARTAO DE CREDITO": "CARTAO_CREDITO",
    "CARTAODECREDITO": "CARTAO_CREDITO",
    "CARTAO DEBITO": "CARTAO_DEBITO",
    "CARTAODEBITO": "CARTAO_DEBITO",
    "CARTAO_DEBITO": "CARTAO_DEBITO",
    "DEBITO": "CARTAO_DEBITO",
    "CARTAO DE DEBITO": "CARTAO_DEBITO",
    "CARTAODEDEBITO": "CARTAO_DEBITO",
    "BOLETO": "BOLETO",
    "OUTRO": "OUTRO",
    "OUTROS": "OUTRO",
}


def _decimal(valor: Any) -> Decimal:
    if valor is None:
        return Decimal("0.00")

    if isinstance(valor, Decimal):
        return valor

    try:
        return Decimal(str(valor))
    except Exception:
        return Decimal("0.00")


def _float(valor: Any) -> float:
    return float(_decimal(valor).quantize(Decimal("0.01")))


def _limpar_texto(valor: Any) -> str:
    if valor is None:
        return ""

    texto = str(valor).strip()
    texto = unicodedata.normalize("NFD", texto)
    texto = "".join(char for char in texto if unicodedata.category(char) != "Mn")
    texto = texto.upper()
    texto = texto.replace("-", " ").replace("_", " ")
    texto = re.sub(r"\s+", " ", texto).strip()
    return texto


def _compactar_texto(valor: Any) -> str:
    return re.sub(r"[^A-Z0-9]", "", _limpar_texto(valor))


def _normalizar_forma_pagamento(valor: Any) -> tuple[str, str]:
    texto = _limpar_texto(valor)
    compacto = _compactar_texto(valor)
    chave = MAPA_FORMAS_PAGAMENTO.get(texto) or MAPA_FORMAS_PAGAMENTO.get(compacto) or "OUTRO"

    for item_chave, item_label in FORMAS_PAGAMENTO_PADRAO:
        if item_chave == chave:
            return item_chave, item_label

    return "OUTRO", "Outro"


def _obter_primeiro_atributo(objeto: Any, nomes: list[str], padrao: Any = None) -> Any:
    for nome in nomes:
        if hasattr(objeto, nome):
            valor = getattr(objeto, nome)
            if valor not in (None, ""):
                return valor

    return padrao


def _forma_pagamento_conta(conta: Any) -> tuple[str, str]:
    valor = _obter_primeiro_atributo(
        conta,
        [
            "forma_pagamento_baixa",
            "forma_pagamento",
            "meio_pagamento",
            "metodo_pagamento",
            "tipo_pagamento",
            "pagamento_forma",
        ],
        "OUTRO",
    )
    return _normalizar_forma_pagamento(valor)


def _observacao_baixa(conta: Any) -> str | None:
    valor = _obter_primeiro_atributo(
        conta,
        [
            "observacao_baixa",
            "observacao_pagamento",
            "observacao",
        ],
        None,
    )
    return str(valor).strip() if valor not in (None, "") else None


def _status_atual(conta: Any) -> str:
    status = _obter_primeiro_atributo(conta, ["status_atual", "status"], "PENDENTE")
    return str(status or "PENDENTE").upper()


def _valor_realizado(conta: Any) -> Decimal:
    valor_pago = _decimal(getattr(conta, "valor_pago", None))
    if valor_pago > 0:
        return valor_pago

    return _decimal(getattr(conta, "valor", None))


def _valor_previsto(conta: Any) -> Decimal:
    return _decimal(getattr(conta, "valor", None))


def _periodo_padrao(data_inicio: date | None, data_fim: date | None) -> tuple[date, date]:
    hoje = date.today()
    inicio = data_inicio or hoje.replace(day=1)
    fim = data_fim or hoje

    if fim < inicio:
        raise HTTPException(
            status_code=400,
            detail="A data final não pode ser menor que a data inicial.",
        )

    return inicio, fim


def _nome_cliente(conta: Any) -> str | None:
    cliente = getattr(conta, "cliente", None)
    if not cliente:
        return None

    for campo in ["nome", "nome_completo", "razao_social", "nome_fantasia"]:
        valor = getattr(cliente, campo, None)
        if valor:
            return str(valor)

    return None


def _nome_fornecedor(conta: Any) -> str | None:
    valor = getattr(conta, "fornecedor", None)
    return str(valor).strip() if valor not in (None, "") else None


def _descricao_conta(conta: Any) -> str:
    descricao = getattr(conta, "descricao", None)
    if descricao:
        return str(descricao)

    origem_tipo = getattr(conta, "origem_tipo", None)
    origem_id = getattr(conta, "origem_id", None)

    if origem_tipo and origem_id:
        return f"{origem_tipo} #{origem_id}"

    return "Lançamento financeiro"


def _classificacao_dre(conta: Any) -> dict[str, str | None]:
    classificacao = getattr(conta, "classificacao_dre", None)

    grupo = getattr(conta, "grupo_dre", None)
    categoria = getattr(conta, "categoria_dre", None)
    subcategoria = getattr(conta, "subcategoria_dre", None)

    if classificacao:
        grupo = grupo or getattr(classificacao, "grupo", None)
        categoria = categoria or getattr(classificacao, "categoria", None)
        subcategoria = subcategoria or getattr(classificacao, "subcategoria", None)

    return {
        "grupo": grupo,
        "categoria": categoria,
        "subcategoria": subcategoria,
    }


def _query_receber_base(db: Session, empresa_id: int):
    return (
        db.query(FinanceiroReceber)
        .options(joinedload(FinanceiroReceber.cliente))
        .filter(FinanceiroReceber.empresa_id == empresa_id)
    )


def _query_pagar_base(db: Session, empresa_id: int):
    return (
        db.query(FinanceiroPagar)
        .options(joinedload(FinanceiroPagar.classificacao_dre))
        .filter(FinanceiroPagar.empresa_id == empresa_id)
    )


def _consultar_recebimentos_realizados(
    db: Session,
    empresa_id: int,
    data_inicio: date,
    data_fim: date,
) -> list[FinanceiroReceber]:
    return (
        _query_receber_base(db, empresa_id)
        .filter(FinanceiroReceber.status == "PAGO")
        .filter(FinanceiroReceber.data_pagamento >= data_inicio)
        .filter(FinanceiroReceber.data_pagamento <= data_fim)
        .order_by(FinanceiroReceber.data_pagamento.asc(), FinanceiroReceber.id.asc())
        .all()
    )


def _consultar_pagamentos_realizados(
    db: Session,
    empresa_id: int,
    data_inicio: date,
    data_fim: date,
) -> list[FinanceiroPagar]:
    return (
        _query_pagar_base(db, empresa_id)
        .filter(FinanceiroPagar.status == "PAGO")
        .filter(FinanceiroPagar.data_pagamento >= data_inicio)
        .filter(FinanceiroPagar.data_pagamento <= data_fim)
        .order_by(FinanceiroPagar.data_pagamento.asc(), FinanceiroPagar.id.asc())
        .all()
    )


def _consultar_recebimentos_previstos(
    db: Session,
    empresa_id: int,
    data_inicio: date,
    data_fim: date,
) -> list[FinanceiroReceber]:
    return (
        _query_receber_base(db, empresa_id)
        .filter(FinanceiroReceber.status.notin_(["PAGO", "CANCELADO"]))
        .filter(FinanceiroReceber.vencimento >= data_inicio)
        .filter(FinanceiroReceber.vencimento <= data_fim)
        .order_by(FinanceiroReceber.vencimento.asc(), FinanceiroReceber.id.asc())
        .all()
    )


def _consultar_pagamentos_previstos(
    db: Session,
    empresa_id: int,
    data_inicio: date,
    data_fim: date,
) -> list[FinanceiroPagar]:
    return (
        _query_pagar_base(db, empresa_id)
        .filter(FinanceiroPagar.status.notin_(["PAGO", "CANCELADO"]))
        .filter(FinanceiroPagar.vencimento >= data_inicio)
        .filter(FinanceiroPagar.vencimento <= data_fim)
        .order_by(FinanceiroPagar.vencimento.asc(), FinanceiroPagar.id.asc())
        .all()
    )


def _consultar_recebimentos_vencidos(
    db: Session,
    empresa_id: int,
    data_base: date,
) -> list[FinanceiroReceber]:
    return (
        _query_receber_base(db, empresa_id)
        .filter(FinanceiroReceber.status.notin_(["PAGO", "CANCELADO"]))
        .filter(FinanceiroReceber.vencimento < data_base)
        .order_by(FinanceiroReceber.vencimento.asc(), FinanceiroReceber.id.asc())
        .all()
    )


def _consultar_pagamentos_vencidos(
    db: Session,
    empresa_id: int,
    data_base: date,
) -> list[FinanceiroPagar]:
    return (
        _query_pagar_base(db, empresa_id)
        .filter(FinanceiroPagar.status.notin_(["PAGO", "CANCELADO"]))
        .filter(FinanceiroPagar.vencimento < data_base)
        .order_by(FinanceiroPagar.vencimento.asc(), FinanceiroPagar.id.asc())
        .all()
    )


def _montar_cards_formas_pagamento(
    recebimentos: list[FinanceiroReceber],
    pagamentos: list[FinanceiroPagar],
) -> list[dict[str, Any]]:
    acumulado: dict[str, dict[str, Any]] = {}

    for chave, label in FORMAS_PAGAMENTO_PADRAO:
        acumulado[chave] = {
            "chave": chave,
            "forma_pagamento": label,
            "entradas": Decimal("0.00"),
            "saidas": Decimal("0.00"),
            "saldo": Decimal("0.00"),
            "quantidade_entradas": 0,
            "quantidade_saidas": 0,
            "quantidade_total": 0,
        }

    for conta in recebimentos:
        chave, label = _forma_pagamento_conta(conta)
        valor = _valor_realizado(conta)

        if chave not in acumulado:
            acumulado[chave] = {
                "chave": chave,
                "forma_pagamento": label,
                "entradas": Decimal("0.00"),
                "saidas": Decimal("0.00"),
                "saldo": Decimal("0.00"),
                "quantidade_entradas": 0,
                "quantidade_saidas": 0,
                "quantidade_total": 0,
            }

        acumulado[chave]["entradas"] += valor
        acumulado[chave]["quantidade_entradas"] += 1
        acumulado[chave]["quantidade_total"] += 1

    for conta in pagamentos:
        chave, label = _forma_pagamento_conta(conta)
        valor = _valor_realizado(conta)

        if chave not in acumulado:
            acumulado[chave] = {
                "chave": chave,
                "forma_pagamento": label,
                "entradas": Decimal("0.00"),
                "saidas": Decimal("0.00"),
                "saldo": Decimal("0.00"),
                "quantidade_entradas": 0,
                "quantidade_saidas": 0,
                "quantidade_total": 0,
            }

        acumulado[chave]["saidas"] += valor
        acumulado[chave]["quantidade_saidas"] += 1
        acumulado[chave]["quantidade_total"] += 1

    resultado = []

    for chave, _label in FORMAS_PAGAMENTO_PADRAO:
        item = acumulado[chave]
        saldo = item["entradas"] - item["saidas"]

        resultado.append(
            {
                "chave": item["chave"],
                "forma_pagamento": item["forma_pagamento"],
                "entradas": _float(item["entradas"]),
                "saidas": _float(item["saidas"]),
                "saldo": _float(saldo),
                "quantidade_entradas": item["quantidade_entradas"],
                "quantidade_saidas": item["quantidade_saidas"],
                "quantidade_total": item["quantidade_total"],
            }
        )

    return resultado


def _movimento_receber(
    conta: FinanceiroReceber,
    natureza: str,
    data_movimento: date,
) -> dict[str, Any]:
    chave, label = _forma_pagamento_conta(conta)
    valor = _valor_realizado(conta) if natureza == "REALIZADO" else _valor_previsto(conta)

    return {
        "id": conta.id,
        "origem": "financeiro_receber",
        "tipo": "ENTRADA",
        "natureza": natureza,
        "data": data_movimento.isoformat(),
        "descricao": _descricao_conta(conta),
        "pessoa": _nome_cliente(conta),
        "forma_pagamento_chave": chave,
        "forma_pagamento": label,
        "entrada": _float(valor),
        "saida": 0.0,
        "valor": _float(valor),
        "status": _status_atual(conta),
        "observacao": getattr(conta, "observacao", None),
        "observacao_baixa": _observacao_baixa(conta),
    }


def _movimento_pagar(
    conta: FinanceiroPagar,
    natureza: str,
    data_movimento: date,
) -> dict[str, Any]:
    chave, label = _forma_pagamento_conta(conta)
    valor = _valor_realizado(conta) if natureza == "REALIZADO" else _valor_previsto(conta)
    dre = _classificacao_dre(conta)

    return {
        "id": conta.id,
        "origem": "financeiro_pagar",
        "tipo": "SAIDA",
        "natureza": natureza,
        "data": data_movimento.isoformat(),
        "descricao": _descricao_conta(conta),
        "pessoa": _nome_fornecedor(conta),
        "forma_pagamento_chave": chave,
        "forma_pagamento": label,
        "entrada": 0.0,
        "saida": _float(valor),
        "valor": _float(valor),
        "status": _status_atual(conta),
        "grupo_dre": dre["grupo"],
        "categoria_dre": dre["categoria"],
        "subcategoria_dre": dre["subcategoria"],
        "observacao": getattr(conta, "observacao", None),
        "observacao_baixa": _observacao_baixa(conta),
    }


def _gerar_movimentos(
    recebimentos_realizados: list[FinanceiroReceber],
    pagamentos_realizados: list[FinanceiroPagar],
    recebimentos_previstos: list[FinanceiroReceber],
    pagamentos_previstos: list[FinanceiroPagar],
    saldo_inicial: Decimal = Decimal("0.00"),
) -> list[dict[str, Any]]:
    movimentos: list[dict[str, Any]] = []

    for conta in recebimentos_realizados:
        if conta.data_pagamento:
            movimentos.append(_movimento_receber(conta, "REALIZADO", conta.data_pagamento))

    for conta in pagamentos_realizados:
        if conta.data_pagamento:
            movimentos.append(_movimento_pagar(conta, "REALIZADO", conta.data_pagamento))

    for conta in recebimentos_previstos:
        if conta.vencimento:
            movimentos.append(_movimento_receber(conta, "PREVISTO", conta.vencimento))

    for conta in pagamentos_previstos:
        if conta.vencimento:
            movimentos.append(_movimento_pagar(conta, "PREVISTO", conta.vencimento))

    prioridade_natureza = {"REALIZADO": 0, "PREVISTO": 1}
    prioridade_tipo = {"ENTRADA": 0, "SAIDA": 1}

    movimentos.sort(
        key=lambda item: (
            item["data"],
            prioridade_natureza.get(item["natureza"], 9),
            prioridade_tipo.get(item["tipo"], 9),
            item["descricao"] or "",
        )
    )

    saldo = saldo_inicial

    for item in movimentos:
        saldo += _decimal(item["entrada"])
        saldo -= _decimal(item["saida"])
        item["saldo_acumulado"] = _float(saldo)

    return movimentos


def _serie_diaria(
    movimentos: list[dict[str, Any]],
    saldo_inicial: Decimal = Decimal("0.00"),
) -> list[dict[str, Any]]:
    dias: dict[str, dict[str, Any]] = {}

    for item in movimentos:
        data_item = item["data"]

        if data_item not in dias:
            dias[data_item] = {
                "data": data_item,
                "entradas": Decimal("0.00"),
                "saidas": Decimal("0.00"),
                "saldo_dia": Decimal("0.00"),
                "saldo_acumulado": Decimal("0.00"),
                "quantidade_movimentos": 0,
            }

        dias[data_item]["entradas"] += _decimal(item["entrada"])
        dias[data_item]["saidas"] += _decimal(item["saida"])
        dias[data_item]["quantidade_movimentos"] += 1

    saldo = saldo_inicial
    resultado = []

    for data_item in sorted(dias.keys()):
        dia = dias[data_item]
        saldo_dia = dia["entradas"] - dia["saidas"]
        saldo += saldo_dia

        resultado.append(
            {
                "data": data_item,
                "entradas": _float(dia["entradas"]),
                "saidas": _float(dia["saidas"]),
                "saldo_dia": _float(saldo_dia),
                "saldo_acumulado": _float(saldo),
                "quantidade_movimentos": dia["quantidade_movimentos"],
            }
        )

    return resultado


def _previsao_periodos(db: Session, empresa_id: int) -> list[dict[str, Any]]:
    hoje = date.today()
    periodos = [7, 15, 30]
    previsoes = []

    for dias in periodos:
        data_fim = hoje + timedelta(days=dias)

        recebimentos = _consultar_recebimentos_previstos(db, empresa_id, hoje, data_fim)
        pagamentos = _consultar_pagamentos_previstos(db, empresa_id, hoje, data_fim)

        entradas = sum((_valor_previsto(conta) for conta in recebimentos), Decimal("0.00"))
        saidas = sum((_valor_previsto(conta) for conta in pagamentos), Decimal("0.00"))
        saldo = entradas - saidas

        previsoes.append(
            {
                "dias": dias,
                "titulo": f"Próximos {dias} dias",
                "data_inicio": hoje.isoformat(),
                "data_fim": data_fim.isoformat(),
                "entradas_previstas": _float(entradas),
                "saidas_previstas": _float(saidas),
                "saldo_previsto": _float(saldo),
                "quantidade_entradas": len(recebimentos),
                "quantidade_saidas": len(pagamentos),
                "quantidade_total": len(recebimentos) + len(pagamentos),
            }
        )

    return previsoes


def _calcular_semaforo(
    saldo_periodo: Decimal,
    previsoes: list[dict[str, Any]],
    saidas_vencidas: Decimal,
    entradas_vencidas: Decimal,
) -> dict[str, Any]:
    previsao_7 = Decimal("0.00")
    previsao_15 = Decimal("0.00")
    previsao_30 = Decimal("0.00")

    for item in previsoes:
        dias = int(item["dias"])
        saldo = _decimal(item["saldo_previsto"])

        if dias == 7:
            previsao_7 = saldo
        elif dias == 15:
            previsao_15 = saldo
        elif dias == 30:
            previsao_30 = saldo

    saldo_projetado_7 = saldo_periodo + previsao_7
    saldo_projetado_15 = saldo_periodo + previsao_15
    saldo_projetado_30 = saldo_periodo + previsao_30

    if saldo_projetado_7 < 0 or saldo_projetado_30 < 0 or (
        saidas_vencidas > 0 and saldo_periodo < saidas_vencidas
    ):
        return {
            "nivel": "VERMELHO",
            "cor": "vermelho",
            "titulo": "Risco de saldo negativo",
            "mensagem": "As saídas vencidas ou previstas podem comprometer o saldo projetado.",
            "saldo_periodo": _float(saldo_periodo),
            "saldo_projetado_7": _float(saldo_projetado_7),
            "saldo_projetado_15": _float(saldo_projetado_15),
            "saldo_projetado_30": _float(saldo_projetado_30),
            "entradas_vencidas": _float(entradas_vencidas),
            "saidas_vencidas": _float(saidas_vencidas),
        }

    if saldo_periodo <= 0 or saldo_projetado_15 < 0 or saidas_vencidas > 0:
        return {
            "nivel": "AMARELO",
            "cor": "amarelo",
            "titulo": "Atenção",
            "mensagem": "O caixa exige acompanhamento próximo nos próximos dias.",
            "saldo_periodo": _float(saldo_periodo),
            "saldo_projetado_7": _float(saldo_projetado_7),
            "saldo_projetado_15": _float(saldo_projetado_15),
            "saldo_projetado_30": _float(saldo_projetado_30),
            "entradas_vencidas": _float(entradas_vencidas),
            "saidas_vencidas": _float(saidas_vencidas),
        }

    return {
        "nivel": "VERDE",
        "cor": "verde",
        "titulo": "Caixa saudável",
        "mensagem": "O saldo realizado e a previsão indicam cenário financeiro controlado.",
        "saldo_periodo": _float(saldo_periodo),
        "saldo_projetado_7": _float(saldo_projetado_7),
        "saldo_projetado_15": _float(saldo_projetado_15),
        "saldo_projetado_30": _float(saldo_projetado_30),
        "entradas_vencidas": _float(entradas_vencidas),
        "saidas_vencidas": _float(saidas_vencidas),
    }


def _montar_dashboard(
    db: Session,
    empresa_id: int,
    data_inicio: date,
    data_fim: date,
) -> dict[str, Any]:
    recebimentos_realizados = _consultar_recebimentos_realizados(
        db,
        empresa_id,
        data_inicio,
        data_fim,
    )
    pagamentos_realizados = _consultar_pagamentos_realizados(
        db,
        empresa_id,
        data_inicio,
        data_fim,
    )
    recebimentos_previstos = _consultar_recebimentos_previstos(
        db,
        empresa_id,
        data_inicio,
        data_fim,
    )
    pagamentos_previstos = _consultar_pagamentos_previstos(
        db,
        empresa_id,
        data_inicio,
        data_fim,
    )

    entradas_realizadas = sum(
        (_valor_realizado(conta) for conta in recebimentos_realizados),
        Decimal("0.00"),
    )
    saidas_realizadas = sum(
        (_valor_realizado(conta) for conta in pagamentos_realizados),
        Decimal("0.00"),
    )
    saldo_realizado = entradas_realizadas - saidas_realizadas

    entradas_previstas = sum(
        (_valor_previsto(conta) for conta in recebimentos_previstos),
        Decimal("0.00"),
    )
    saidas_previstas = sum(
        (_valor_previsto(conta) for conta in pagamentos_previstos),
        Decimal("0.00"),
    )
    saldo_previsto = entradas_previstas - saidas_previstas

    hoje = date.today()
    recebimentos_vencidos = _consultar_recebimentos_vencidos(db, empresa_id, hoje)
    pagamentos_vencidos = _consultar_pagamentos_vencidos(db, empresa_id, hoje)

    entradas_vencidas = sum(
        (_valor_previsto(conta) for conta in recebimentos_vencidos),
        Decimal("0.00"),
    )
    saidas_vencidas = sum(
        (_valor_previsto(conta) for conta in pagamentos_vencidos),
        Decimal("0.00"),
    )

    previsoes = _previsao_periodos(db, empresa_id)

    semaforo = _calcular_semaforo(
        saldo_realizado,
        previsoes,
        saidas_vencidas,
        entradas_vencidas,
    )

    movimentos = _gerar_movimentos(
        recebimentos_realizados=recebimentos_realizados,
        pagamentos_realizados=pagamentos_realizados,
        recebimentos_previstos=recebimentos_previstos,
        pagamentos_previstos=pagamentos_previstos,
    )

    return {
        "empresa_id": empresa_id,
        "data_inicio": data_inicio.isoformat(),
        "data_fim": data_fim.isoformat(),
        "resumo": {
            "entradas": _float(entradas_realizadas),
            "saidas": _float(saidas_realizadas),
            "saldo": _float(saldo_realizado),
            "entradas_realizadas": _float(entradas_realizadas),
            "saidas_realizadas": _float(saidas_realizadas),
            "saldo_realizado": _float(saldo_realizado),
            "entradas_previstas": _float(entradas_previstas),
            "saidas_previstas": _float(saidas_previstas),
            "saldo_previsto": _float(saldo_previsto),
            "entradas_vencidas": _float(entradas_vencidas),
            "saidas_vencidas": _float(saidas_vencidas),
            "quantidade_entradas_realizadas": len(recebimentos_realizados),
            "quantidade_saidas_realizadas": len(pagamentos_realizados),
            "quantidade_entradas_previstas": len(recebimentos_previstos),
            "quantidade_saidas_previstas": len(pagamentos_previstos),
            "quantidade_entradas_vencidas": len(recebimentos_vencidos),
            "quantidade_saidas_vencidas": len(pagamentos_vencidos),
        },
        "formas_pagamento": _montar_cards_formas_pagamento(
            recebimentos_realizados,
            pagamentos_realizados,
        ),
        "previsoes": previsoes,
        "semaforo": semaforo,
        "serie_diaria": _serie_diaria(movimentos),
    }


def _gerar_dados_exportacao(
    db: Session,
    empresa_id: int,
    data_inicio: date,
    data_fim: date,
    incluir_previstos: bool,
    saldo_inicial: Decimal,
) -> dict[str, Any]:
    recebimentos_realizados = _consultar_recebimentos_realizados(db, empresa_id, data_inicio, data_fim)
    pagamentos_realizados = _consultar_pagamentos_realizados(db, empresa_id, data_inicio, data_fim)

    recebimentos_previstos: list[FinanceiroReceber] = []
    pagamentos_previstos: list[FinanceiroPagar] = []

    if incluir_previstos:
        recebimentos_previstos = _consultar_recebimentos_previstos(db, empresa_id, data_inicio, data_fim)
        pagamentos_previstos = _consultar_pagamentos_previstos(db, empresa_id, data_inicio, data_fim)

    movimentos = _gerar_movimentos(
        recebimentos_realizados=recebimentos_realizados,
        pagamentos_realizados=pagamentos_realizados,
        recebimentos_previstos=recebimentos_previstos,
        pagamentos_previstos=pagamentos_previstos,
        saldo_inicial=saldo_inicial,
    )

    entradas = sum((_decimal(item["entrada"]) for item in movimentos), Decimal("0.00"))
    saidas = sum((_decimal(item["saida"]) for item in movimentos), Decimal("0.00"))
    saldo_final = saldo_inicial + entradas - saidas

    return {
        "empresa_id": empresa_id,
        "data_inicio": data_inicio,
        "data_fim": data_fim,
        "incluir_previstos": incluir_previstos,
        "saldo_inicial": saldo_inicial,
        "saldo_final": saldo_final,
        "entradas": entradas,
        "saidas": saidas,
        "movimentos": movimentos,
        "serie_diaria": _serie_diaria(movimentos, saldo_inicial),
        "formas_pagamento": _montar_cards_formas_pagamento(
            recebimentos_realizados,
            pagamentos_realizados,
        ),
    }


def _formatar_data_excel(valor: Any) -> str:
    if not valor:
        return ""

    if isinstance(valor, date):
        return valor.strftime("%d/%m/%Y")

    try:
        partes = str(valor).split("-")
        if len(partes) == 3:
            return f"{partes[2]}/{partes[1]}/{partes[0]}"
    except Exception:
        pass

    return str(valor)


def _aplicar_estilo_cabecalho(ws, linha: int, coluna_inicial: int, coluna_final: int) -> None:
    fill = PatternFill("solid", fgColor="0F172A")
    font = Font(color="FFFFFF", bold=True)
    alignment = Alignment(horizontal="center", vertical="center")
    border = Border(bottom=Side(style="thin", color="CBD5E1"))

    for col in range(coluna_inicial, coluna_final + 1):
        cell = ws.cell(row=linha, column=col)
        cell.fill = fill
        cell.font = font
        cell.alignment = alignment
        cell.border = border


def _aplicar_bordas(ws, linha_inicial: int, linha_final: int, coluna_inicial: int, coluna_final: int) -> None:
    border = Border(
        left=Side(style="thin", color="E2E8F0"),
        right=Side(style="thin", color="E2E8F0"),
        top=Side(style="thin", color="E2E8F0"),
        bottom=Side(style="thin", color="E2E8F0"),
    )

    for row in range(linha_inicial, linha_final + 1):
        for col in range(coluna_inicial, coluna_final + 1):
            ws.cell(row=row, column=col).border = border


def _ajustar_larguras(ws) -> None:
    for column_cells in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column_cells[0].column)

        for cell in column_cells:
            valor = cell.value
            if valor is None:
                continue

            max_length = max(max_length, len(str(valor)))

        ws.column_dimensions[column_letter].width = min(max(max_length + 3, 12), 42)


def _formatar_moeda(ws, colunas: list[int], linha_inicial: int, linha_final: int) -> None:
    for row in range(linha_inicial, linha_final + 1):
        for col in colunas:
            ws.cell(row=row, column=col).number_format = 'R$ #,##0.00;[Red]-R$ #,##0.00'


def _criar_xlsx_fluxo_caixa(dados: dict[str, Any]) -> BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "Fluxo Caixa"

    ws["A1"] = "Fluxo de Caixa"
    ws["A1"].font = Font(size=18, bold=True, color="0F172A")
    ws["A2"] = f"Empresa: {dados['empresa_id']}"
    ws["A3"] = f"Período: {_formatar_data_excel(dados['data_inicio'])} a {_formatar_data_excel(dados['data_fim'])}"
    ws["A4"] = f"Gerado em: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    ws["A5"] = "Inclui previstos: Sim" if dados["incluir_previstos"] else "Inclui previstos: Não"

    resumo_linha = 7
    ws.cell(row=resumo_linha, column=1, value="Indicador")
    ws.cell(row=resumo_linha, column=2, value="Valor")
    _aplicar_estilo_cabecalho(ws, resumo_linha, 1, 2)

    resumo = [
        ("Saldo inicial", _float(dados["saldo_inicial"])),
        ("Entradas", _float(dados["entradas"])),
        ("Saídas", _float(dados["saidas"])),
        ("Saldo final", _float(dados["saldo_final"])),
        ("Quantidade de movimentos", len(dados["movimentos"])),
    ]

    for index, item in enumerate(resumo, start=resumo_linha + 1):
        ws.cell(row=index, column=1, value=item[0])
        ws.cell(row=index, column=2, value=item[1])

    _formatar_moeda(ws, [2], resumo_linha + 1, resumo_linha + 4)
    _aplicar_bordas(ws, resumo_linha, resumo_linha + len(resumo), 1, 2)

    linha_movimentos = resumo_linha + len(resumo) + 3
    headers = [
        "Data",
        "Tipo",
        "Natureza",
        "Descrição",
        "Pessoa",
        "Forma de pagamento",
        "Status",
        "Grupo DRE",
        "Categoria DRE",
        "Subcategoria DRE",
        "Observação da baixa",
        "Entrada",
        "Saída",
        "Saldo acumulado",
    ]

    for col, titulo in enumerate(headers, start=1):
        ws.cell(row=linha_movimentos, column=col, value=titulo)

    _aplicar_estilo_cabecalho(ws, linha_movimentos, 1, len(headers))

    for row_index, item in enumerate(dados["movimentos"], start=linha_movimentos + 1):
        ws.cell(row=row_index, column=1, value=_formatar_data_excel(item.get("data")))
        ws.cell(row=row_index, column=2, value=item.get("tipo"))
        ws.cell(row=row_index, column=3, value=item.get("natureza"))
        ws.cell(row=row_index, column=4, value=item.get("descricao"))
        ws.cell(row=row_index, column=5, value=item.get("pessoa"))
        ws.cell(row=row_index, column=6, value=item.get("forma_pagamento"))
        ws.cell(row=row_index, column=7, value=item.get("status"))
        ws.cell(row=row_index, column=8, value=item.get("grupo_dre"))
        ws.cell(row=row_index, column=9, value=item.get("categoria_dre"))
        ws.cell(row=row_index, column=10, value=item.get("subcategoria_dre"))
        ws.cell(row=row_index, column=11, value=item.get("observacao_baixa"))
        ws.cell(row=row_index, column=12, value=item.get("entrada"))
        ws.cell(row=row_index, column=13, value=item.get("saida"))
        ws.cell(row=row_index, column=14, value=item.get("saldo_acumulado"))

    ultima_linha_movimentos = max(linha_movimentos + 1, linha_movimentos + len(dados["movimentos"]))
    _formatar_moeda(ws, [12, 13, 14], linha_movimentos + 1, ultima_linha_movimentos)
    _aplicar_bordas(ws, linha_movimentos, ultima_linha_movimentos, 1, len(headers))

    for row in ws.iter_rows(min_row=linha_movimentos + 1, max_row=ultima_linha_movimentos):
        tipo = row[1].value
        natureza = row[2].value

        if tipo == "ENTRADA":
            row[1].font = Font(color="15803D", bold=True)
        elif tipo == "SAIDA":
            row[1].font = Font(color="B91C1C", bold=True)

        if natureza == "PREVISTO":
            for cell in row:
                cell.fill = PatternFill("solid", fgColor="F8FAFC")

    ws.freeze_panes = f"A{linha_movimentos + 1}"
    ws.auto_filter.ref = f"A{linha_movimentos}:N{ultima_linha_movimentos}"

    ws_diario = wb.create_sheet("Diario")
    ws_diario["A1"] = "Tabela diária"
    ws_diario["A1"].font = Font(size=16, bold=True, color="0F172A")

    headers_diario = [
        "Data",
        "Entradas",
        "Saídas",
        "Saldo do dia",
        "Saldo acumulado",
        "Movimentos",
    ]

    for col, titulo in enumerate(headers_diario, start=1):
        ws_diario.cell(row=3, column=col, value=titulo)

    _aplicar_estilo_cabecalho(ws_diario, 3, 1, len(headers_diario))

    for row_index, item in enumerate(dados["serie_diaria"], start=4):
        ws_diario.cell(row=row_index, column=1, value=_formatar_data_excel(item.get("data")))
        ws_diario.cell(row=row_index, column=2, value=item.get("entradas"))
        ws_diario.cell(row=row_index, column=3, value=item.get("saidas"))
        ws_diario.cell(row=row_index, column=4, value=item.get("saldo_dia"))
        ws_diario.cell(row=row_index, column=5, value=item.get("saldo_acumulado"))
        ws_diario.cell(row=row_index, column=6, value=item.get("quantidade_movimentos"))

    ultima_linha_diario = max(4, 3 + len(dados["serie_diaria"]))
    _formatar_moeda(ws_diario, [2, 3, 4, 5], 4, ultima_linha_diario)
    _aplicar_bordas(ws_diario, 3, ultima_linha_diario, 1, len(headers_diario))
    ws_diario.freeze_panes = "A4"
    ws_diario.auto_filter.ref = f"A3:F{ultima_linha_diario}"

    ws_formas = wb.create_sheet("Formas Pagamento")
    ws_formas["A1"] = "Formas de pagamento"
    ws_formas["A1"].font = Font(size=16, bold=True, color="0F172A")

    headers_formas = [
        "Forma de pagamento",
        "Entradas",
        "Saídas",
        "Saldo",
        "Qtd. entradas",
        "Qtd. saídas",
        "Qtd. total",
    ]

    for col, titulo in enumerate(headers_formas, start=1):
        ws_formas.cell(row=3, column=col, value=titulo)

    _aplicar_estilo_cabecalho(ws_formas, 3, 1, len(headers_formas))

    for row_index, item in enumerate(dados["formas_pagamento"], start=4):
        ws_formas.cell(row=row_index, column=1, value=item.get("forma_pagamento"))
        ws_formas.cell(row=row_index, column=2, value=item.get("entradas"))
        ws_formas.cell(row=row_index, column=3, value=item.get("saidas"))
        ws_formas.cell(row=row_index, column=4, value=item.get("saldo"))
        ws_formas.cell(row=row_index, column=5, value=item.get("quantidade_entradas"))
        ws_formas.cell(row=row_index, column=6, value=item.get("quantidade_saidas"))
        ws_formas.cell(row=row_index, column=7, value=item.get("quantidade_total"))

    ultima_linha_formas = max(4, 3 + len(dados["formas_pagamento"]))
    _formatar_moeda(ws_formas, [2, 3, 4], 4, ultima_linha_formas)
    _aplicar_bordas(ws_formas, 3, ultima_linha_formas, 1, len(headers_formas))
    ws_formas.freeze_panes = "A4"
    ws_formas.auto_filter.ref = f"A3:G{ultima_linha_formas}"

    for sheet in wb.worksheets:
        _ajustar_larguras(sheet)

        for row in sheet.iter_rows():
            for cell in row:
                cell.alignment = Alignment(
                    vertical="center",
                    horizontal=cell.alignment.horizontal or "left",
                    wrap_text=True,
                )

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


@router.get("/resumo")
def resumo_fluxo_caixa(
    empresa_id: int = Query(..., ge=1),
    data_inicio: date | None = Query(None),
    data_fim: date | None = Query(None),
    db: Session = Depends(get_db),
):
    inicio, fim = _periodo_padrao(data_inicio, data_fim)
    return _montar_dashboard(db, empresa_id, inicio, fim)


@router.get("/dashboard")
def dashboard_fluxo_caixa(
    empresa_id: int = Query(..., ge=1),
    data_inicio: date | None = Query(None),
    data_fim: date | None = Query(None),
    db: Session = Depends(get_db),
):
    inicio, fim = _periodo_padrao(data_inicio, data_fim)
    return _montar_dashboard(db, empresa_id, inicio, fim)


@router.get("/formas-pagamento")
def formas_pagamento_fluxo_caixa(
    empresa_id: int = Query(..., ge=1),
    data_inicio: date | None = Query(None),
    data_fim: date | None = Query(None),
    db: Session = Depends(get_db),
):
    inicio, fim = _periodo_padrao(data_inicio, data_fim)

    recebimentos = _consultar_recebimentos_realizados(db, empresa_id, inicio, fim)
    pagamentos = _consultar_pagamentos_realizados(db, empresa_id, inicio, fim)

    return {
        "empresa_id": empresa_id,
        "data_inicio": inicio.isoformat(),
        "data_fim": fim.isoformat(),
        "formas_pagamento": _montar_cards_formas_pagamento(recebimentos, pagamentos),
    }


@router.get("/previsao")
def previsao_fluxo_caixa(
    empresa_id: int = Query(..., ge=1),
    db: Session = Depends(get_db),
):
    hoje = date.today()
    previsoes = _previsao_periodos(db, empresa_id)

    recebimentos_vencidos = _consultar_recebimentos_vencidos(db, empresa_id, hoje)
    pagamentos_vencidos = _consultar_pagamentos_vencidos(db, empresa_id, hoje)

    entradas_vencidas = sum(
        (_valor_previsto(conta) for conta in recebimentos_vencidos),
        Decimal("0.00"),
    )
    saidas_vencidas = sum(
        (_valor_previsto(conta) for conta in pagamentos_vencidos),
        Decimal("0.00"),
    )

    semaforo = _calcular_semaforo(
        Decimal("0.00"),
        previsoes,
        saidas_vencidas,
        entradas_vencidas,
    )

    return {
        "empresa_id": empresa_id,
        "data_base": hoje.isoformat(),
        "previsoes": previsoes,
        "semaforo": semaforo,
    }


@router.get("/extrato")
def extrato_fluxo_caixa(
    empresa_id: int = Query(..., ge=1),
    data_inicio: date | None = Query(None),
    data_fim: date | None = Query(None),
    incluir_previstos: bool = Query(True),
    saldo_inicial: Decimal = Query(Decimal("0.00")),
    db: Session = Depends(get_db),
):
    inicio, fim = _periodo_padrao(data_inicio, data_fim)

    dados = _gerar_dados_exportacao(
        db=db,
        empresa_id=empresa_id,
        data_inicio=inicio,
        data_fim=fim,
        incluir_previstos=incluir_previstos,
        saldo_inicial=saldo_inicial,
    )

    movimentos = dados["movimentos"]
    entradas = dados["entradas"]
    saidas = dados["saidas"]
    saldo_final = dados["saldo_final"]

    return {
        "empresa_id": empresa_id,
        "data_inicio": inicio.isoformat(),
        "data_fim": fim.isoformat(),
        "incluir_previstos": incluir_previstos,
        "saldo_inicial": _float(saldo_inicial),
        "saldo_final": _float(saldo_final),
        "resumo": {
            "entradas": _float(entradas),
            "saidas": _float(saidas),
            "saldo": _float(entradas - saidas),
            "quantidade_movimentos": len(movimentos),
        },
        "movimentos": movimentos,
    }


@router.get("/serie-diaria")
def serie_diaria_fluxo_caixa(
    empresa_id: int = Query(..., ge=1),
    data_inicio: date | None = Query(None),
    data_fim: date | None = Query(None),
    incluir_previstos: bool = Query(True),
    saldo_inicial: Decimal = Query(Decimal("0.00")),
    db: Session = Depends(get_db),
):
    inicio, fim = _periodo_padrao(data_inicio, data_fim)

    dados = _gerar_dados_exportacao(
        db=db,
        empresa_id=empresa_id,
        data_inicio=inicio,
        data_fim=fim,
        incluir_previstos=incluir_previstos,
        saldo_inicial=saldo_inicial,
    )

    return {
        "empresa_id": empresa_id,
        "data_inicio": inicio.isoformat(),
        "data_fim": fim.isoformat(),
        "incluir_previstos": incluir_previstos,
        "saldo_inicial": _float(saldo_inicial),
        "serie_diaria": dados["serie_diaria"],
    }


@router.get("/exportar-xlsx")
def exportar_xlsx_fluxo_caixa(
    empresa_id: int = Query(..., ge=1),
    data_inicio: date | None = Query(None),
    data_fim: date | None = Query(None),
    incluir_previstos: bool = Query(True),
    saldo_inicial: Decimal = Query(Decimal("0.00")),
    db: Session = Depends(get_db),
):
    inicio, fim = _periodo_padrao(data_inicio, data_fim)

    dados = _gerar_dados_exportacao(
        db=db,
        empresa_id=empresa_id,
        data_inicio=inicio,
        data_fim=fim,
        incluir_previstos=incluir_previstos,
        saldo_inicial=saldo_inicial,
    )

    arquivo = _criar_xlsx_fluxo_caixa(dados)
    nome_arquivo = f"fluxo_caixa_empresa_{empresa_id}_{inicio.isoformat()}_{fim.isoformat()}.xlsx"

    return StreamingResponse(
        arquivo,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f'attachment; filename="{nome_arquivo}"'
        },
    )
