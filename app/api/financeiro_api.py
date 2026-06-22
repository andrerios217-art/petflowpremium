from datetime import date, datetime
from io import BytesIO
from calendar import monthrange
from decimal import Decimal

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from sqlalchemy import extract, text
from sqlalchemy.orm import Session
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from app.core.deps import get_db
from app.models.financeiro_receber import FinanceiroReceber
from app.schemas.financeiro import (
    FinanceiroReceberBaixa,
    FinanceiroReceberCreate,
    FinanceiroReceberListOut,
    FinanceiroReceberOut,
)

router = APIRouter(prefix="/api/financeiro", tags=["Financeiro"])


def _serializar(conta: FinanceiroReceber):
    return {
        "id": conta.id,
        "empresa_id": conta.empresa_id,
        "cliente_id": conta.cliente_id,
        "cliente_nome": conta.cliente.nome if conta.cliente else None,
        "origem_tipo": conta.origem_tipo,
        "origem_id": conta.origem_id,
        "descricao": conta.descricao,
        "observacao": conta.observacao,
        "valor": float(conta.valor or 0),
        "valor_pago": float(conta.valor_pago or 0),
        "vencimento": conta.vencimento.isoformat() if conta.vencimento else None,
        "data_pagamento": conta.data_pagamento.isoformat() if conta.data_pagamento else None,
        "status": conta.status,
        "status_atual": conta.status_atual,
        "esta_vencido": conta.esta_vencido,
        "created_at": conta.created_at.isoformat() if conta.created_at else None,
        "updated_at": conta.updated_at.isoformat() if conta.updated_at else None,
    }


@router.post("/receber", response_model=FinanceiroReceberOut)
def criar_conta_receber(
    payload: FinanceiroReceberCreate,
    db: Session = Depends(get_db),
):
    conta = FinanceiroReceber(
        empresa_id=payload.empresa_id,
        cliente_id=payload.cliente_id,
        origem_tipo=payload.origem_tipo,
        origem_id=payload.origem_id,
        descricao=payload.descricao,
        observacao=payload.observacao,
        valor=Decimal(payload.valor),
        valor_pago=Decimal("0.00"),
        vencimento=payload.vencimento,
        status="PENDENTE",
    )

    db.add(conta)
    db.commit()
    db.refresh(conta)

    return _serializar(conta)


@router.get("/receber", response_model=FinanceiroReceberListOut)
def listar_contas_receber(
    empresa_id: int = Query(..., ge=1),
    mes: int | None = Query(None, ge=1, le=12),
    ano: int | None = Query(None, ge=2000, le=2100),
    db: Session = Depends(get_db),
):
    query = db.query(FinanceiroReceber).filter(FinanceiroReceber.empresa_id == empresa_id)

    if ano is not None:
        query = query.filter(extract("year", FinanceiroReceber.vencimento) == ano)

    if mes is not None:
        query = query.filter(extract("month", FinanceiroReceber.vencimento) == mes)

    contas = (
        query
        .order_by(FinanceiroReceber.vencimento.asc(), FinanceiroReceber.id.desc())
        .all()
    )

    hoje = date.today()
    total_pendente = Decimal("0")
    total_pago = Decimal("0")
    total_vencido = Decimal("0")
    qtd_pendente = 0
    qtd_pago = 0
    qtd_vencido = 0
    resultado = []

    for conta in contas:
        status = conta.status

        if status != "PAGO" and conta.vencimento < hoje:
            status = "VENCIDO"

        valor = Decimal(conta.valor or 0)

        if status == "PAGO":
            total_pago += valor
            qtd_pago += 1
        elif status == "VENCIDO":
            total_vencido += valor
            qtd_vencido += 1
        else:
            total_pendente += valor
            qtd_pendente += 1

        resultado.append(_serializar(conta))

    return {
        "empresa_id": empresa_id,
        "resumo": {
            "total_pendente": float(total_pendente),
            "total_pago": float(total_pago),
            "total_vencido": float(total_vencido),
            "quantidade_pendente": qtd_pendente,
            "quantidade_paga": qtd_pago,
            "quantidade_vencida": qtd_vencido,
        },
        "contas": resultado,
    }


@router.post("/receber/{conta_id}/baixar")
def baixar_conta_receber(
    conta_id: int,
    payload: FinanceiroReceberBaixa,
    db: Session = Depends(get_db),
):
    conta = db.query(FinanceiroReceber).filter(FinanceiroReceber.id == conta_id).first()

    if not conta:
        raise HTTPException(status_code=404, detail="Conta não encontrada.")

    if conta.status == "PAGO":
        raise HTTPException(status_code=400, detail="Esta conta já está paga.")

    conta.status = "PAGO"
    conta.data_pagamento = payload.data_pagamento or date.today()
    conta.valor_pago = Decimal(payload.valor_pago or conta.valor or 0)

    db.commit()

    return {"ok": True}

def _parse_competencia_inicio(valor: str | None) -> date | None:
    if not valor:
        return None
    try:
        ano, mes = valor.split("-")
        return date(int(ano), int(mes), 1)
    except Exception:
        raise HTTPException(status_code=400, detail="Competencia inicial invalida. Use YYYY-MM.")


def _parse_competencia_fim(valor: str | None) -> date | None:
    if not valor:
        return None
    try:
        ano, mes = valor.split("-")
        ultimo_dia = monthrange(int(ano), int(mes))[1]
        return date(int(ano), int(mes), ultimo_dia)
    except Exception:
        raise HTTPException(status_code=400, detail="Competencia final invalida. Use YYYY-MM.")


def _moeda_excel(cell):
    cell.number_format = 'R$ #,##0.00;[Red]-R$ #,##0.00'


def _data_excel(cell):
    cell.number_format = "dd/mm/yyyy"


def _aplicar_estilo_tabela(ws, header_row: int, max_row: int, max_col: int):
    header_fill = PatternFill("solid", fgColor="1F2937")
    header_font = Font(color="FFFFFF", bold=True)
    zebra_fill = PatternFill("solid", fgColor="F8FAFC")
    border = Border(
        left=Side(style="thin", color="E5E7EB"),
        right=Side(style="thin", color="E5E7EB"),
        top=Side(style="thin", color="E5E7EB"),
        bottom=Side(style="thin", color="E5E7EB"),
    )

    for cell in ws[header_row]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    for row in range(header_row + 1, max_row + 1):
        for col in range(1, max_col + 1):
            cell = ws.cell(row=row, column=col)
            cell.border = border
            cell.alignment = Alignment(vertical="center")
            if row % 2 == 0:
                cell.fill = zebra_fill

    ws.auto_filter.ref = f"A{header_row}:{get_column_letter(max_col)}{max_row}"
    ws.freeze_panes = f"A{header_row + 1}"


def _ajustar_colunas(ws, larguras: dict[int, int]):
    for col_idx, width in larguras.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def _titulo_aba(ws, titulo: str, subtitulo: str, col_final: int):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=col_final)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=col_final)

    ws["A1"] = titulo
    ws["A2"] = subtitulo

    ws["A1"].font = Font(size=16, bold=True, color="1F2937")
    ws["A2"].font = Font(size=10, color="64748B")

    ws.row_dimensions[1].height = 24
    ws.row_dimensions[2].height = 20


def _dre_txt(value: str) -> str:
    return value


def _dre_parse_inicio(valor: str | None) -> date:
    if not valor:
        hoje = date.today()
        return date(hoje.year, 1, 1)

    try:
        ano, mes = valor.split("-")
        return date(int(ano), int(mes), 1)
    except Exception:
        raise HTTPException(status_code=400, detail="Competencia inicial invalida. Use YYYY-MM.")


def _dre_parse_fim(valor: str | None) -> date:
    if not valor:
        hoje = date.today()
        return date(hoje.year, hoje.month, monthrange(hoje.year, hoje.month)[1])

    try:
        ano, mes = valor.split("-")
        return date(int(ano), int(mes), monthrange(int(ano), int(mes))[1])
    except Exception:
        raise HTTPException(status_code=400, detail="Competencia final invalida. Use YYYY-MM.")


def _dre_months_between(inicio: date, fim: date) -> list[date]:
    meses = []
    atual = date(inicio.year, inicio.month, 1)

    while atual <= fim:
        meses.append(atual)

        if atual.month == 12:
            atual = date(atual.year + 1, 1, 1)
        else:
            atual = date(atual.year, atual.month + 1, 1)

    return meses


def _dre_month_key(valor) -> str:
    if not valor:
        return ""
    return f"{valor.year:04d}-{valor.month:02d}"


def _dre_month_label(valor: date) -> str:
    return f"{valor.month:02d}/{valor.year}"


def _dre_money(cell):
    cell.number_format = 'R$ #,##0.00;[Red]-R$ #,##0.00;R$ 0.00'


def _dre_date(cell):
    cell.number_format = "dd/mm/yyyy"


def _dre_header_style(cell):
    cell.fill = PatternFill("solid", fgColor="1F2937")
    cell.font = Font(color="FFFFFF", bold=True)
    cell.alignment = Alignment(horizontal="center", vertical="center")


def _dre_title(ws, title: str, subtitle: str, end_col: int):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=end_col)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=end_col)

    ws["A1"] = title
    ws["A2"] = subtitle

    ws["A1"].font = Font(size=16, bold=True, color="111827")
    ws["A2"].font = Font(size=10, color="64748B")

    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")

    ws.row_dimensions[1].height = 26
    ws.row_dimensions[2].height = 22


def _dre_apply_table_style(ws, header_row: int, max_row: int, max_col: int):
    thin = Side(style="thin", color="E5E7EB")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    zebra = PatternFill("solid", fgColor="F8FAFC")

    for col in range(1, max_col + 1):
        cell = ws.cell(row=header_row, column=col)
        _dre_header_style(cell)
        cell.border = border

    for row in range(header_row + 1, max_row + 1):
        for col in range(1, max_col + 1):
            cell = ws.cell(row=row, column=col)
            cell.border = border
            cell.alignment = Alignment(vertical="center")
            if row % 2 == 0:
                cell.fill = zebra

    ws.freeze_panes = ws.cell(row=header_row + 1, column=2).coordinate
    ws.auto_filter.ref = f"A{header_row}:{get_column_letter(max_col)}{max_row}"


def _dre_widths(ws, widths: dict[int, int]):
    for col, width in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = width


def _dre_style_matrix_row(ws, row: int, level: int, kind: str, first_value_col: int, last_col: int):
    fills = {
        "header_receita": "DBEAFE",
        "header_despesa": "FEE2E2",
        "resultado": "DCFCE7",
        "subtotal": "F3F4F6",
        "normal": "FFFFFF",
    }

    font_colors = {
        "header_receita": "1E3A8A",
        "header_despesa": "991B1B",
        "resultado": "166534",
        "subtotal": "111827",
        "normal": "111827",
    }

    fill = PatternFill("solid", fgColor=fills.get(kind, fills["normal"]))
    font = Font(
        bold=kind in ("header_receita", "header_despesa", "resultado", "subtotal"),
        color=font_colors.get(kind, "111827"),
    )

    for col in range(1, last_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(vertical="center")

        if col == 1:
            cell.alignment = Alignment(
                vertical="center",
                horizontal="left",
                indent=max(0, level),
            )

        if col >= first_value_col:
            _dre_money(cell)


def _dre_status_fill(cell, status: str):
    status = str(status or "").upper()

    if status in ("PAGO", "RECEBIDO"):
        cell.fill = PatternFill("solid", fgColor="DCFCE7")
        cell.font = Font(color="166534", bold=True)
    elif status in ("VENCIDO", "ATRASADO"):
        cell.fill = PatternFill("solid", fgColor="FEE2E2")
        cell.font = Font(color="991B1B", bold=True)
    else:
        cell.fill = PatternFill("solid", fgColor="FEF3C7")
        cell.font = Font(color="92400E", bold=True)


@router.get("/dre/conferencia.xlsx")
def exportar_dre_conferencia_xlsx(
    empresa_id: int = Query(..., ge=1),
    competencia_inicio: str | None = Query(None),
    competencia_fim: str | None = Query(None),
    db: Session = Depends(get_db),
):
    data_inicio = _dre_parse_inicio(competencia_inicio)
    data_fim = _dre_parse_fim(competencia_fim)

    if data_fim < data_inicio:
        raise HTTPException(status_code=400, detail="Competencia final nao pode ser menor que a inicial.")

    meses = _dre_months_between(data_inicio, data_fim)
    mes_keys = [_dre_month_key(m) for m in meses]
    mes_labels = [_dre_month_label(m) for m in meses]

    params = {
        "empresa_id": empresa_id,
        "data_inicio": data_inicio,
        "data_fim": data_fim,
    }

    entradas = db.execute(
        text("""
            SELECT
                r.id,
                r.cliente_id,
                c.nome AS cliente_nome,
                r.descricao,
                r.observacao,
                r.valor,
                r.valor_pago,
                r.vencimento,
                r.data_pagamento,
                r.status,
                r.origem_tipo,
                r.origem_id
            FROM financeiro_receber r
            LEFT JOIN clientes c ON c.id = r.cliente_id
            WHERE r.empresa_id = :empresa_id
              AND r.vencimento >= :data_inicio
              AND r.vencimento <= :data_fim
            ORDER BY r.vencimento ASC, r.id ASC
        """),
        params,
    ).mappings().all()

    saidas = db.execute(
        text("""
            SELECT
                p.id,
                p.descricao,
                p.fornecedor,
                p.observacao,
                p.valor,
                p.valor_pago,
                p.vencimento,
                p.data_pagamento,
                p.status,
                p.origem_tipo,
                p.origem_id,
                p.classificacao_dre_id,
                d.grupo,
                d.categoria,
                d.subcategoria
            FROM financeiro_pagar p
            LEFT JOIN financeiro_plano_dre d ON d.id = p.classificacao_dre_id
            WHERE p.empresa_id = :empresa_id
              AND p.vencimento >= :data_inicio
              AND p.vencimento <= :data_fim
            ORDER BY p.vencimento ASC, p.id ASC
        """),
        params,
    ).mappings().all()

    plano = db.execute(
        text("""
            SELECT id, grupo, categoria, subcategoria, ordem, ativo
            FROM financeiro_plano_dre
            WHERE empresa_id = :empresa_id
            ORDER BY ordem ASC, grupo ASC, categoria ASC, subcategoria ASC
        """),
        {"empresa_id": empresa_id},
    ).mappings().all()

    wb = Workbook()
    ws_dre = wb.active
    ws_dre.title = "DRE"
    ws_entradas = wb.create_sheet("Entradas")
    ws_saidas = wb.create_sheet("Saidas")
    ws_conferencia = wb.create_sheet("Conferencia")
    ws_plano = wb.create_sheet("Plano DRE")

    subtitle = (
        f"Empresa {empresa_id} | Periodo {data_inicio.strftime('%d/%m/%Y')} "
        f"a {data_fim.strftime('%d/%m/%Y')} | Gerado em {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    )

    dre_col_total = 2 + len(meses)
    dre_last_col = dre_col_total

    _dre_title(ws_dre, "DRE - Demonstrativo do Resultado", subtitle, dre_last_col)

    headers = ["Nome"] + mes_labels + ["Total"]
    for col, header in enumerate(headers, start=1):
        ws_dre.cell(row=4, column=col, value=header)

    _dre_apply_table_style(ws_dre, 4, 4, dre_last_col)
    _dre_widths(ws_dre, {1: 44, **{i: 15 for i in range(2, dre_last_col + 1)}})
    ws_dre.freeze_panes = "B5"

    receita_por_descricao = {}
    receita_total_meses = {key: 0.0 for key in mes_keys}

    for item in entradas:
        key = _dre_month_key(item["vencimento"])
        descricao = item["descricao"] or "Receitas"
        valor = float(item["valor"] or 0)

        receita_total_meses[key] = receita_total_meses.get(key, 0.0) + valor

        if descricao not in receita_por_descricao:
            receita_por_descricao[descricao] = {k: 0.0 for k in mes_keys}

        receita_por_descricao[descricao][key] = receita_por_descricao[descricao].get(key, 0.0) + valor

    despesa_total_meses = {key: 0.0 for key in mes_keys}
    despesas_tree = {}

    for item in saidas:
        key = _dre_month_key(item["vencimento"])
        grupo = item["grupo"] or "Sem grupo"
        categoria = item["categoria"] or "Sem categoria"
        subcategoria = item["subcategoria"] or item["descricao"] or "Sem subcategoria"
        valor = -abs(float(item["valor"] or 0))

        despesa_total_meses[key] = despesa_total_meses.get(key, 0.0) + valor

        despesas_tree.setdefault(grupo, {})
        despesas_tree[grupo].setdefault(categoria, {})
        despesas_tree[grupo][categoria].setdefault(subcategoria, {k: 0.0 for k in mes_keys})
        despesas_tree[grupo][categoria][subcategoria][key] += valor

    row = 5

    def write_matrix_line(nome, valores_meses, level=0, kind="normal"):
        nonlocal row
        ws_dre.cell(row=row, column=1, value=nome)

        total = 0.0
        for idx, key in enumerate(mes_keys, start=2):
            valor = float(valores_meses.get(key, 0) or 0)
            total += valor
            ws_dre.cell(row=row, column=idx, value=valor)

        ws_dre.cell(row=row, column=dre_col_total, value=total)
        _dre_style_matrix_row(ws_dre, row, level, kind, 2, dre_last_col)
        row += 1

    write_matrix_line("1 - Receitas", receita_total_meses, 0, "header_receita")

    for idx, descricao in enumerate(sorted(receita_por_descricao.keys()), start=1):
        write_matrix_line(f"  1.{idx} - {descricao}", receita_por_descricao[descricao], 1, "normal")

    write_matrix_line("2 - Despesas", despesa_total_meses, 0, "header_despesa")

    grupo_idx = 1
    for grupo in sorted(despesas_tree.keys()):
        grupo_meses = {key: 0.0 for key in mes_keys}

        for categorias in despesas_tree[grupo].values():
            for subcats in categorias.values():
                for key in mes_keys:
                    grupo_meses[key] += subcats.get(key, 0.0)

        write_matrix_line(f"  2.{grupo_idx} - {grupo}", grupo_meses, 1, "subtotal")

        categoria_idx = 1
        for categoria in sorted(despesas_tree[grupo].keys()):
            categoria_meses = {key: 0.0 for key in mes_keys}

            for subcats in despesas_tree[grupo][categoria].values():
                for key in mes_keys:
                    categoria_meses[key] += subcats.get(key, 0.0)

            write_matrix_line(f"    2.{grupo_idx}.{categoria_idx} - {categoria}", categoria_meses, 2, "normal")

            sub_idx = 1
            for subcategoria in sorted(despesas_tree[grupo][categoria].keys()):
                write_matrix_line(
                    f"      2.{grupo_idx}.{categoria_idx}.{sub_idx} - {subcategoria}",
                    despesas_tree[grupo][categoria][subcategoria],
                    3,
                    "normal",
                )
                sub_idx += 1

            categoria_idx += 1

        grupo_idx += 1

    resultado_meses = {
        key: receita_total_meses.get(key, 0.0) + despesa_total_meses.get(key, 0.0)
        for key in mes_keys
    }

    write_matrix_line("3 - Resultado Operacional", resultado_meses, 0, "resultado")
    write_matrix_line("4 - Lucro/Prejuizo Liquido", resultado_meses, 0, "resultado")

    _dre_apply_table_style(ws_dre, 4, row - 1, dre_last_col)

    for r in range(5, row):
        for c in range(2, dre_last_col + 1):
            _dre_money(ws_dre.cell(row=r, column=c))

    ws_dre.auto_filter.ref = f"A4:{get_column_letter(dre_last_col)}{row - 1}"

    _dre_title(ws_entradas, "Entradas financeiras", subtitle, 12)
    entradas_header = [
        "ID", "Cliente ID", "Cliente", "Descricao", "Observacao", "Valor",
        "Valor pago", "Vencimento", "Data pagamento", "Status", "Origem tipo", "Origem ID"
    ]

    for col, header in enumerate(entradas_header, start=1):
        ws_entradas.cell(row=4, column=col, value=header)

    for row_idx, item in enumerate(entradas, start=5):
        values = [
            item["id"],
            item["cliente_id"],
            item["cliente_nome"],
            item["descricao"],
            item["observacao"],
            float(item["valor"] or 0),
            float(item["valor_pago"] or 0),
            item["vencimento"],
            item["data_pagamento"],
            item["status"],
            item["origem_tipo"],
            item["origem_id"],
        ]

        for col, value in enumerate(values, start=1):
            ws_entradas.cell(row=row_idx, column=col, value=value)

        _dre_money(ws_entradas.cell(row=row_idx, column=6))
        _dre_money(ws_entradas.cell(row=row_idx, column=7))
        _dre_date(ws_entradas.cell(row=row_idx, column=8))
        _dre_date(ws_entradas.cell(row=row_idx, column=9))
        _dre_status_fill(ws_entradas.cell(row=row_idx, column=10), item["status"])

    _dre_apply_table_style(ws_entradas, 4, max(4, ws_entradas.max_row), len(entradas_header))
    _dre_widths(ws_entradas, {1: 8, 2: 12, 3: 26, 4: 34, 5: 36, 6: 16, 7: 16, 8: 14, 9: 16, 10: 14, 11: 16, 12: 12})

    _dre_title(ws_saidas, "Saidas financeiras", subtitle, 16)
    saidas_header = [
        "ID", "Descricao", "Fornecedor", "Observacao", "Valor", "Valor pago",
        "Vencimento", "Data pagamento", "Status", "Grupo DRE", "Categoria DRE",
        "Subcategoria DRE", "Origem tipo", "Origem ID", "Classificacao DRE ID", "Competencia"
    ]

    for col, header in enumerate(saidas_header, start=1):
        ws_saidas.cell(row=4, column=col, value=header)

    for row_idx, item in enumerate(saidas, start=5):
        competencia = _dre_month_label(item["vencimento"]) if item["vencimento"] else ""

        values = [
            item["id"],
            item["descricao"],
            item["fornecedor"],
            item["observacao"],
            float(item["valor"] or 0),
            float(item["valor_pago"] or 0),
            item["vencimento"],
            item["data_pagamento"],
            item["status"],
            item["grupo"] or "Sem grupo",
            item["categoria"] or "Sem categoria",
            item["subcategoria"] or "Sem subcategoria",
            item["origem_tipo"],
            item["origem_id"],
            item["classificacao_dre_id"],
            competencia,
        ]

        for col, value in enumerate(values, start=1):
            ws_saidas.cell(row=row_idx, column=col, value=value)

        _dre_money(ws_saidas.cell(row=row_idx, column=5))
        _dre_money(ws_saidas.cell(row=row_idx, column=6))
        _dre_date(ws_saidas.cell(row=row_idx, column=7))
        _dre_date(ws_saidas.cell(row=row_idx, column=8))
        _dre_status_fill(ws_saidas.cell(row=row_idx, column=9), item["status"])

    _dre_apply_table_style(ws_saidas, 4, max(4, ws_saidas.max_row), len(saidas_header))
    _dre_widths(ws_saidas, {1: 8, 2: 34, 3: 24, 4: 36, 5: 16, 6: 16, 7: 14, 8: 16, 9: 14, 10: 24, 11: 24, 12: 28, 13: 14, 14: 12, 15: 18, 16: 14})

    _dre_title(ws_conferencia, "Conferencia consolidada", subtitle, 7)
    conferencia_header = ["Tipo", "Grupo", "Categoria", "Subcategoria", "Previsto", "Realizado", "Diferenca"]

    for col, header in enumerate(conferencia_header, start=1):
        ws_conferencia.cell(row=4, column=col, value=header)

    resumo = {}

    for item in entradas:
        chave = ("Entrada", "Receitas", "Receitas", item["descricao"] or "Entradas")
        resumo.setdefault(chave, [0.0, 0.0])
        resumo[chave][0] += float(item["valor"] or 0)
        resumo[chave][1] += float(item["valor_pago"] or 0)

    for item in saidas:
        chave = (
            "Saida",
            item["grupo"] or "Sem grupo",
            item["categoria"] or "Sem categoria",
            item["subcategoria"] or "Sem subcategoria",
        )
        resumo.setdefault(chave, [0.0, 0.0])
        resumo[chave][0] += float(item["valor"] or 0)
        resumo[chave][1] += float(item["valor_pago"] or 0)

    for row_idx, (chave, valores) in enumerate(sorted(resumo.items()), start=5):
        previsto, realizado = valores

        ws_conferencia.cell(row=row_idx, column=1, value=chave[0])
        ws_conferencia.cell(row=row_idx, column=2, value=chave[1])
        ws_conferencia.cell(row=row_idx, column=3, value=chave[2])
        ws_conferencia.cell(row=row_idx, column=4, value=chave[3])
        ws_conferencia.cell(row=row_idx, column=5, value=previsto)
        ws_conferencia.cell(row=row_idx, column=6, value=realizado)
        ws_conferencia.cell(row=row_idx, column=7, value=realizado - previsto)

        _dre_money(ws_conferencia.cell(row=row_idx, column=5))
        _dre_money(ws_conferencia.cell(row=row_idx, column=6))
        _dre_money(ws_conferencia.cell(row=row_idx, column=7))

    _dre_apply_table_style(ws_conferencia, 4, max(4, ws_conferencia.max_row), len(conferencia_header))
    _dre_widths(ws_conferencia, {1: 14, 2: 26, 3: 26, 4: 34, 5: 16, 6: 16, 7: 16})

    _dre_title(ws_plano, "Plano DRE", subtitle, 6)
    plano_header = ["ID", "Grupo", "Categoria", "Subcategoria", "Ordem", "Ativo"]

    for col, header in enumerate(plano_header, start=1):
        ws_plano.cell(row=4, column=col, value=header)

    for row_idx, item in enumerate(plano, start=5):
        values = [
            item["id"],
            item["grupo"],
            item["categoria"],
            item["subcategoria"],
            item["ordem"],
            "Sim" if item["ativo"] else "Nao",
        ]

        for col, value in enumerate(values, start=1):
            ws_plano.cell(row=row_idx, column=col, value=value)

    _dre_apply_table_style(ws_plano, 4, max(4, ws_plano.max_row), len(plano_header))
    _dre_widths(ws_plano, {1: 8, 2: 28, 3: 28, 4: 34, 5: 10, 6: 10})

    for ws in wb.worksheets:
        ws.sheet_view.showGridLines = False

        for row_cells in ws.iter_rows():
            for cell in row_cells:
                cell.font = cell.font.copy(name="Arial")
                cell.alignment = cell.alignment.copy(vertical="center")

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = (
        f"dre-conferencia-empresa-{empresa_id}-"
        f"{data_inicio.strftime('%Y-%m')}-a-{data_fim.strftime('%Y-%m')}.xlsx"
    )

    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )

