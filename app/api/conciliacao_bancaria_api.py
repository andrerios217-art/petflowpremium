from datetime import date, datetime, timedelta
from decimal import Decimal
from io import BytesIO, StringIO
from typing import Any
import csv
import re
import unicodedata

from fastapi import APIRouter, Body, Depends, File, HTTPException, Query, UploadFile
from fastapi.responses import StreamingResponse
from sqlalchemy import text
from sqlalchemy.orm import Session, joinedload

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from app.core.deps import get_db
from app.models.financeiro_pagar import FinanceiroPagar
from app.models.financeiro_receber import FinanceiroReceber
from app.models.conciliacao_bancaria import ConciliacaoBancaria


router = APIRouter(prefix="/api/conciliacao-bancaria", tags=["Conciliação Bancária"])


CABECALHOS_DATA = [
    "data",
    "dt",
    "data movimento",
    "data do movimento",
    "data lancamento",
    "data lançamento",
    "lancamento",
    "lançamento",
]

CABECALHOS_DESCRICAO = [
    "descricao",
    "descrição",
    "historico",
    "histórico",
    "lancamento",
    "lançamento",
    "movimento",
    "memo",
    "detalhe",
    "detalhes",
    "nome",
    "favorecido",
    "beneficiario",
    "beneficiário",
    "pagador",
    "recebedor",
    "razao social",
    "razão social",
]

CABECALHOS_VALOR = [
    "valor",
    "vlr",
    "amount",
    "valor lancamento",
    "valor lançamento",
    "valor movimento",
]

CABECALHOS_ENTRADA = [
    "entrada",
    "credito",
    "crédito",
    "credit",
    "receita",
    "recebimento",
]

CABECALHOS_SAIDA = [
    "saida",
    "saída",
    "debito",
    "débito",
    "debit",
    "despesa",
    "pagamento",
]

CABECALHOS_TIPO = [
    "tipo",
    "natureza",
    "operacao",
    "operação",
    "dc",
    "d/c",
    "credito debito",
    "crédito débito",
]

CABECALHOS_DOCUMENTO = [
    "documento",
    "doc",
    "id",
    "identificador",
    "numero",
    "número",
    "nsu",
]

PALAVRAS_ENTRADA = [
    "credito",
    "crédito",
    "entrada",
    "receita",
    "recebimento",
    "deposito",
    "depósito",
    "pix recebido",
    "c",
]

PALAVRAS_SAIDA = [
    "debito",
    "débito",
    "saida",
    "saída",
    "despesa",
    "pagamento",
    "pix enviado",
    "tarifa",
    "d",
]


def _limpar_texto(valor: Any) -> str:
    if valor is None:
        return ""

    texto = str(valor).strip()
    texto = unicodedata.normalize("NFD", texto)
    texto = "".join(char for char in texto if unicodedata.category(char) != "Mn")
    texto = texto.lower()
    texto = texto.replace("_", " ").replace("-", " ")
    texto = re.sub(r"\s+", " ", texto).strip()
    return texto


def _compactar_texto(valor: Any) -> str:
    return re.sub(r"[^a-z0-9]", "", _limpar_texto(valor))


def _texto_para_busca(valor: Any) -> str:
    texto = _limpar_texto(valor)
    texto = re.sub(r"[^a-z0-9\s]", " ", texto)
    texto = re.sub(r"\s+", " ", texto).strip()
    return texto


def _decimal(valor: Any) -> Decimal:
    if valor is None:
        return Decimal("0.00")

    if isinstance(valor, Decimal):
        return valor

    texto = str(valor).strip()

    if not texto:
        return Decimal("0.00")

    negativo = False

    if texto.startswith("(") and texto.endswith(")"):
        negativo = True
        texto = texto[1:-1]

    texto = texto.replace("R$", "")
    texto = texto.replace("r$", "")
    texto = texto.replace(" ", "")

    if texto.startswith("-"):
        negativo = True
        texto = texto[1:]

    texto = re.sub(r"[^0-9,.-]", "", texto)

    if "," in texto and "." in texto:
        if texto.rfind(",") > texto.rfind("."):
            texto = texto.replace(".", "").replace(",", ".")
        else:
            texto = texto.replace(",", "")
    elif "," in texto:
        texto = texto.replace(".", "").replace(",", ".")

    try:
        valor_decimal = Decimal(texto or "0")
    except Exception:
        valor_decimal = Decimal("0.00")

    return -valor_decimal if negativo else valor_decimal


def _float(valor: Any) -> float:
    return float(_decimal(valor).quantize(Decimal("0.01")))


def _parse_data(valor: Any) -> date | None:
    if valor is None:
        return None

    texto = str(valor).strip()

    if not texto:
        return None

    texto = texto.split(" ")[0].strip()

    formatos = [
        "%d/%m/%Y",
        "%d/%m/%y",
        "%Y-%m-%d",
        "%d-%m-%Y",
        "%d-%m-%y",
        "%Y/%m/%d",
    ]

    for formato in formatos:
        try:
            return datetime.strptime(texto, formato).date()
        except Exception:
            pass

    return None


def _detectar_coluna(cabecalhos: list[str], candidatos: list[str]) -> str | None:
    mapa = {_compactar_texto(cabecalho): cabecalho for cabecalho in cabecalhos}

    for candidato in candidatos:
        chave = _compactar_texto(candidato)
        if chave in mapa:
            return mapa[chave]

    for cabecalho in cabecalhos:
        cabecalho_limpo = _limpar_texto(cabecalho)
        for candidato in candidatos:
            candidato_limpo = _limpar_texto(candidato)
            if candidato_limpo and candidato_limpo in cabecalho_limpo:
                return cabecalho

    return None


def _status_atual(conta: Any) -> str:
    status = getattr(conta, "status", None) or getattr(conta, "status_atual", None) or "PENDENTE"
    return str(status).upper()


def _valor_realizado(conta: Any) -> Decimal:
    valor_pago = _decimal(getattr(conta, "valor_pago", None))

    if valor_pago > 0:
        return valor_pago

    return _decimal(getattr(conta, "valor", None))


def _descricao_conta(conta: Any) -> str:
    descricao = getattr(conta, "descricao", None)

    if descricao:
        return str(descricao)

    origem_tipo = getattr(conta, "origem_tipo", None)
    origem_id = getattr(conta, "origem_id", None)

    if origem_tipo and origem_id:
        return f"{origem_tipo} #{origem_id}"

    return "Lançamento financeiro"


def _nome_cliente(conta: Any) -> str | None:
    cliente = getattr(conta, "cliente", None)

    if not cliente:
        return None

    for campo in ["nome", "nome_completo", "razao_social", "nome_fantasia"]:
        valor = getattr(cliente, campo, None)
        if valor:
            return str(valor)

    return None


def _nome_fornecedor(conta: Any) -> str | None:
    valor = getattr(conta, "fornecedor", None)
    return str(valor).strip() if valor not in (None, "") else None


def _forma_pagamento(conta: Any) -> str | None:
    for campo in [
        "forma_pagamento_baixa",
        "forma_pagamento",
        "meio_pagamento",
        "metodo_pagamento",
        "tipo_pagamento",
        "pagamento_forma",
    ]:
        if hasattr(conta, campo):
            valor = getattr(conta, campo)
            if valor not in (None, ""):
                return str(valor)

    return None


def _linha_financeiro_receber(conta: FinanceiroReceber) -> dict[str, Any]:
    return {
        "origem": "financeiro_receber",
        "id": conta.id,
        "tipo": "ENTRADA",
        "data": conta.data_pagamento,
        "data_iso": conta.data_pagamento.isoformat() if conta.data_pagamento else None,
        "descricao": _descricao_conta(conta),
        "pessoa": _nome_cliente(conta),
        "forma_pagamento": _forma_pagamento(conta),
        "valor": _valor_realizado(conta),
        "valor_float": _float(_valor_realizado(conta)),
        "status": _status_atual(conta),
        "texto_busca": _texto_para_busca(f"{_descricao_conta(conta)} {_nome_cliente(conta) or ''}"),
    }


def _linha_financeiro_pagar(conta: FinanceiroPagar) -> dict[str, Any]:
    return {
        "origem": "financeiro_pagar",
        "id": conta.id,
        "tipo": "SAIDA",
        "data": conta.data_pagamento,
        "data_iso": conta.data_pagamento.isoformat() if conta.data_pagamento else None,
        "descricao": _descricao_conta(conta),
        "pessoa": _nome_fornecedor(conta),
        "forma_pagamento": _forma_pagamento(conta),
        "valor": _valor_realizado(conta),
        "valor_float": _float(_valor_realizado(conta)),
        "status": _status_atual(conta),
        "texto_busca": _texto_para_busca(f"{_descricao_conta(conta)} {_nome_fornecedor(conta) or ''}"),
    }


def _consultar_financeiro_realizado(
    db: Session,
    empresa_id: int,
    data_inicio: date,
    data_fim: date,
) -> list[dict[str, Any]]:
    recebimentos = (
        db.query(FinanceiroReceber)
        .options(joinedload(FinanceiroReceber.cliente))
        .filter(FinanceiroReceber.empresa_id == empresa_id)
        .filter(FinanceiroReceber.status == "PAGO")
        .filter(FinanceiroReceber.data_pagamento >= data_inicio)
        .filter(FinanceiroReceber.data_pagamento <= data_fim)
        .order_by(FinanceiroReceber.data_pagamento.asc(), FinanceiroReceber.id.asc())
        .all()
    )

    pagamentos = (
        db.query(FinanceiroPagar)
        .filter(FinanceiroPagar.empresa_id == empresa_id)
        .filter(FinanceiroPagar.status == "PAGO")
        .filter(FinanceiroPagar.data_pagamento >= data_inicio)
        .filter(FinanceiroPagar.data_pagamento <= data_fim)
        .order_by(FinanceiroPagar.data_pagamento.asc(), FinanceiroPagar.id.asc())
        .all()
    )

    movimentos = [_linha_financeiro_receber(item) for item in recebimentos]
    movimentos.extend([_linha_financeiro_pagar(item) for item in pagamentos])

    movimentos.sort(
        key=lambda item: (
            item["data"] or date.min,
            item["tipo"],
            item["descricao"] or "",
        )
    )

    return movimentos


def _ler_upload_texto(arquivo: UploadFile) -> str:
    conteudo = arquivo.file.read()

    if not conteudo:
        raise HTTPException(status_code=400, detail="Arquivo vazio.")

    for encoding in ["utf-8-sig", "utf-8", "cp1252", "latin-1"]:
        try:
            return conteudo.decode(encoding)
        except Exception:
            pass

    raise HTTPException(
        status_code=400,
        detail="Não foi possível ler o arquivo. Envie CSV em UTF-8, ANSI ou Latin-1.",
    )


def _detectar_dialeto_csv(texto: str):
    amostra = texto[:4096]

    try:
        return csv.Sniffer().sniff(amostra, delimiters=";,|\t,")
    except Exception:
        class DialetoFallback(csv.excel):
            delimiter = ";"

        return DialetoFallback


def _descricao_csv_por_fallback(
    linha: list[str],
    cabecalhos: list[str],
    coluna_data: str | None,
    coluna_valor: str | None,
    coluna_entrada: str | None,
    coluna_saida: str | None,
    coluna_tipo: str | None,
    coluna_documento: str | None,
) -> str:
    ignorar = {
        _limpar_texto(coluna_data or ""),
        _limpar_texto(coluna_valor or ""),
        _limpar_texto(coluna_entrada or ""),
        _limpar_texto(coluna_saida or ""),
        _limpar_texto(coluna_tipo or ""),
        _limpar_texto(coluna_documento or ""),
        "saldo",
        "saldo r",
        "saldo rs",
        "valor",
        "valor r",
        "valor rs",
        "data",
        "cpf cnpj",
        "cpfcnpj",
    }

    partes_prioritarias: list[str] = []
    partes_secundarias: list[str] = []

    for indice, valor in enumerate(linha):
        texto = str(valor or "").strip()

        if not texto:
            continue

        cabecalho = cabecalhos[indice] if indice < len(cabecalhos) else ""
        cabecalho_limpo = _limpar_texto(cabecalho)

        if cabecalho_limpo in ignorar:
            continue

        if _parse_data(texto):
            continue

        if _decimal(texto) != Decimal("0.00"):
            continue

        texto_normalizado = " ".join(texto.split())

        if cabecalho_limpo in [
            "lancamento",
            "lançamento",
            "historico",
            "histórico",
            "descricao",
            "descrição",
            "movimento",
        ]:
            partes_prioritarias.append(texto_normalizado)
            continue

        if cabecalho_limpo in [
            "razao social",
            "razão social",
            "nome",
            "favorecido",
            "beneficiario",
            "beneficiário",
            "pagador",
            "recebedor",
        ]:
            partes_secundarias.append(texto_normalizado)
            continue

        # Fallback prático: em extratos, a segunda coluna costuma ser o lançamento.
        if indice == 1:
            partes_prioritarias.append(texto_normalizado)

    partes = []

    for texto in partes_prioritarias + partes_secundarias:
        if texto and texto not in partes:
            partes.append(texto)

    return " - ".join(partes).strip()


def _parse_csv(texto: str) -> list[dict[str, Any]]:
    import unicodedata

    texto = texto.replace("\ufeff", "").strip()

    if not texto:
        raise HTTPException(status_code=400, detail="CSV vazio.")

    def normalizar(valor: Any) -> str:
        texto_local = str(valor or "").strip().lower()
        texto_local = unicodedata.normalize("NFD", texto_local)
        texto_local = "".join(ch for ch in texto_local if unicodedata.category(ch) != "Mn")
        texto_local = texto_local.replace("(", " ").replace(")", " ")
        texto_local = texto_local.replace("/", " ").replace("\\", " ").replace("-", " ")
        texto_local = texto_local.replace("$", " ").replace(".", " ")
        texto_local = " ".join(texto_local.split())
        return texto_local

    def decimal_csv(valor: Any) -> Decimal:
        raw = str(valor or "").strip()

        if not raw:
            return Decimal("0.00")

        raw = raw.replace("R$", "")
        raw = raw.replace("\xa0", "")
        raw = raw.replace(" ", "")

        negativo = False

        if raw.startswith("(") and raw.endswith(")"):
            negativo = True
            raw = raw[1:-1]

        if raw.startswith("-"):
            negativo = True
            raw = raw[1:]

        raw = raw.replace("+", "")

        if "," in raw:
            raw = raw.replace(".", "").replace(",", ".")
        else:
            raw = raw.replace(",", "")

        try:
            numero = Decimal(raw or "0")
        except Exception:
            return Decimal("0.00")

        return -numero if negativo else numero

    def valor_celula(linha: list[str], indice: int | None) -> str:
        if indice is None:
            return ""

        if indice < 0 or indice >= len(linha):
            return ""

        return str(linha[indice] or "").strip()

    dialeto = _detectar_dialeto_csv(texto)
    reader = csv.reader(StringIO(texto), dialect=dialeto)
    linhas_brutas = [linha for linha in reader if any(str(campo).strip() for campo in linha)]

    if not linhas_brutas:
        raise HTTPException(status_code=400, detail="CSV vazio.")

    indice_cabecalho = None
    cabecalhos: list[str] = []

    for indice, linha in enumerate(linhas_brutas):
        normalizados = [normalizar(campo) for campo in linha]

        tem_data = any(campo == "data" or campo.startswith("data ") for campo in normalizados)
        tem_valor = any("valor" in campo and "saldo" not in campo for campo in normalizados)

        if tem_data and tem_valor:
            indice_cabecalho = indice
            cabecalhos = [str(campo).strip() for campo in linha]
            break

    if indice_cabecalho is None:
        raise HTTPException(
            status_code=400,
            detail="Não encontrei a linha de cabeçalho do extrato. O CSV precisa ter colunas como Data, Lançamento e Valor.",
        )

    cabecalhos_normalizados = [normalizar(campo) for campo in cabecalhos]

    def achar_coluna(*termos: str, contem: bool = False, excluir: list[str] | None = None) -> int | None:
        excluir = excluir or []
        termos_norm = [normalizar(termo) for termo in termos]
        excluir_norm = [normalizar(termo) for termo in excluir]

        for posicao, cabecalho in enumerate(cabecalhos_normalizados):
            if any(item and item in cabecalho for item in excluir_norm):
                continue

            if contem:
                if any(termo and termo in cabecalho for termo in termos_norm):
                    return posicao
            else:
                if cabecalho in termos_norm:
                    return posicao

        return None

    coluna_data = achar_coluna("data", "dt")
    coluna_lancamento = achar_coluna(
        "lancamento",
        "lançamento",
        "historico",
        "histórico",
        "descricao",
        "descrição",
        "movimento",
    )
    coluna_razao_social = achar_coluna(
        "razao social",
        "razão social",
        "nome",
        "favorecido",
        "beneficiario",
        "beneficiário",
        "pagador",
        "recebedor",
    )
    coluna_documento = achar_coluna("cpf cnpj", "cpf/cnpj", "documento", "doc", "cpf", "cnpj")
    coluna_valor = achar_coluna("valor", "valor r", "valor rs", contem=True, excluir=["saldo"])

    if coluna_data is None:
        raise HTTPException(
            status_code=400,
            detail="Não encontrei a coluna de data no CSV.",
        )

    if coluna_valor is None:
        raise HTTPException(
            status_code=400,
            detail="Não encontrei a coluna de valor no CSV.",
        )

    movimentos: list[dict[str, Any]] = []

    for numero_linha, linha in enumerate(linhas_brutas[indice_cabecalho + 1:], start=indice_cabecalho + 2):
        if not any(str(campo).strip() for campo in linha):
            continue

        data_movimento = _parse_data(valor_celula(linha, coluna_data))

        if not data_movimento:
            continue

        valor_original = decimal_csv(valor_celula(linha, coluna_valor))

        if valor_original == 0:
            continue

        lancamento = valor_celula(linha, coluna_lancamento)
        razao_social = valor_celula(linha, coluna_razao_social)
        documento = valor_celula(linha, coluna_documento)

        partes_descricao = []

        if lancamento:
            partes_descricao.append(" ".join(lancamento.split()))

        if razao_social and normalizar(razao_social) not in [normalizar(item) for item in partes_descricao]:
            partes_descricao.append(" ".join(razao_social.split()))

        descricao = " - ".join(partes_descricao).strip()

        if not descricao:
            descricao = valor_celula(linha, 1)

        if not descricao:
            descricao = "Movimento bancário"

        tipo = "ENTRADA" if valor_original > 0 else "SAIDA"
        valor_abs = abs(valor_original)

        movimentos.append(
            {
                "linha": numero_linha,
                "data": data_movimento,
                "data_iso": data_movimento.isoformat(),
                "descricao": descricao,
                "documento": documento or None,
                "tipo": tipo,
                "valor": valor_abs,
                "valor_float": _float(valor_abs),
                "valor_original": _float(valor_original),
                "texto_busca": _texto_para_busca(f"{descricao} {documento}"),
                "conciliado": False,
                "match": None,
                "score": 0,
                "motivo": None,
            }
        )

    if not movimentos:
        raise HTTPException(
            status_code=400,
            detail="Nenhum movimento válido encontrado no CSV. Verifique se as datas e valores estão nas colunas corretas.",
        )

    movimentos.sort(
        key=lambda item: (
            item["data"],
            item["tipo"],
            item["descricao"] or "",
        )
    )

    return movimentos

def _parse_ofx(texto: str) -> list[dict[str, Any]]:
    texto = texto.replace("\r", "\n")
    blocos = re.findall(r"<STMTTRN>(.*?)(?=<STMTTRN>|</BANKTRANLIST>|</CREDITCARDMSGSRSV1>|$)", texto, flags=re.I | re.S)

    movimentos: list[dict[str, Any]] = []

    for numero, bloco in enumerate(blocos, start=1):
        tipo_ofx = _extrair_tag_ofx(bloco, "TRNTYPE") or ""
        data_raw = _extrair_tag_ofx(bloco, "DTPOSTED") or _extrair_tag_ofx(bloco, "DTUSER")
        valor_raw = _extrair_tag_ofx(bloco, "TRNAMT")
        descricao = (
            _extrair_tag_ofx(bloco, "MEMO")
            or _extrair_tag_ofx(bloco, "NAME")
            or _extrair_tag_ofx(bloco, "PAYEE")
            or "Movimento bancário"
        )
        documento = _extrair_tag_ofx(bloco, "FITID") or _extrair_tag_ofx(bloco, "CHECKNUM")

        data_movimento = _parse_data_ofx(data_raw)
        valor = _decimal(valor_raw)

        if not data_movimento or valor == 0:
            continue

        tipo_limpo = _limpar_texto(tipo_ofx)

        if tipo_limpo in ["debit", "debito", "payment", "pagamento"] and valor > 0:
            valor = -valor

        if tipo_limpo in ["credit", "credito", "dep", "deposit", "deposito"] and valor < 0:
            valor = abs(valor)

        tipo = "ENTRADA" if valor > 0 else "SAIDA"

        movimentos.append(
            {
                "linha": numero,
                "data": data_movimento,
                "data_iso": data_movimento.isoformat(),
                "descricao": descricao.strip(),
                "documento": documento.strip() if documento else None,
                "tipo": tipo,
                "valor": abs(valor),
                "valor_float": _float(abs(valor)),
                "valor_original": _float(valor),
                "texto_busca": _texto_para_busca(f"{descricao} {documento or ''}"),
                "conciliado": False,
                "match": None,
                "score": 0,
                "motivo": None,
            }
        )

    if not movimentos:
        raise HTTPException(
            status_code=400,
            detail="Nenhum movimento válido encontrado no OFX.",
        )

    movimentos.sort(
        key=lambda item: (
            item["data"],
            item["tipo"],
            item["descricao"] or "",
        )
    )

    return movimentos


def _extrair_tag_ofx(bloco: str, tag: str) -> str | None:
    padrao_fechado = re.search(rf"<{tag}>(.*?)</{tag}>", bloco, flags=re.I | re.S)

    if padrao_fechado:
        return padrao_fechado.group(1).strip()

    padrao_aberto = re.search(rf"<{tag}>([^\n\r<]+)", bloco, flags=re.I)

    if padrao_aberto:
        return padrao_aberto.group(1).strip()

    return None


def _parse_data_ofx(valor: Any) -> date | None:
    if not valor:
        return None

    texto = str(valor).strip()
    texto = re.sub(r"[^0-9]", "", texto)

    if len(texto) >= 8:
        try:
            return datetime.strptime(texto[:8], "%Y%m%d").date()
        except Exception:
            return None

    return None


def _pontuar_match(
    movimento_banco: dict[str, Any],
    movimento_sistema: dict[str, Any],
    tolerancia_valor: Decimal,
    tolerancia_dias: int,
) -> tuple[int, list[str]]:
    score = 0
    motivos: list[str] = []

    if movimento_banco["tipo"] != movimento_sistema["tipo"]:
        return 0, []

    diferenca_valor = abs(_decimal(movimento_banco["valor"]) - _decimal(movimento_sistema["valor"]))

    if diferenca_valor <= tolerancia_valor:
        score += 55
        motivos.append("valor compatível")
    else:
        return 0, []

    data_banco = movimento_banco["data"]
    data_sistema = movimento_sistema["data"]

    if not data_banco or not data_sistema:
        return 0, []

    diferenca_dias = abs((data_banco - data_sistema).days)

    if diferenca_dias == 0:
        score += 35
        motivos.append("mesma data")
    elif diferenca_dias <= tolerancia_dias:
        score += max(8, 28 - (diferenca_dias * 5))
        motivos.append(f"data próxima ({diferenca_dias} dia(s))")
    else:
        return 0, []

    texto_banco = movimento_banco.get("texto_busca") or ""
    texto_sistema = movimento_sistema.get("texto_busca") or ""

    palavras_banco = set(palavra for palavra in texto_banco.split() if len(palavra) >= 4)
    palavras_sistema = set(palavra for palavra in texto_sistema.split() if len(palavra) >= 4)

    intersecao = palavras_banco.intersection(palavras_sistema)

    if intersecao:
        bonus = min(10, len(intersecao) * 3)
        score += bonus
        motivos.append("descrição semelhante")

    return min(score, 100), motivos


def _conciliar_movimentos(
    movimentos_banco: list[dict[str, Any]],
    movimentos_sistema: list[dict[str, Any]],
    tolerancia_valor: Decimal,
    tolerancia_dias: int,
) -> dict[str, Any]:
    ids_sistema_usados: set[str] = set()
    conciliados = []
    pendentes_banco = []

    for movimento_banco in movimentos_banco:
        melhor_match = None
        melhor_score = 0
        melhor_motivos: list[str] = []

        for movimento_sistema in movimentos_sistema:
            chave_sistema = f"{movimento_sistema['origem']}:{movimento_sistema['id']}"

            if chave_sistema in ids_sistema_usados:
                continue

            score, motivos = _pontuar_match(
                movimento_banco=movimento_banco,
                movimento_sistema=movimento_sistema,
                tolerancia_valor=tolerancia_valor,
                tolerancia_dias=tolerancia_dias,
            )

            if score > melhor_score:
                melhor_score = score
                melhor_match = movimento_sistema
                melhor_motivos = motivos

        item_banco = _serializar_movimento_banco(movimento_banco)

        if melhor_match and melhor_score >= 75:
            chave_sistema = f"{melhor_match['origem']}:{melhor_match['id']}"
            ids_sistema_usados.add(chave_sistema)

            conciliados.append(
                {
                    "banco": item_banco,
                    "sistema": _serializar_movimento_sistema(melhor_match),
                    "score": melhor_score,
                    "motivo": ", ".join(melhor_motivos),
                }
            )
        else:
            sugestoes = _buscar_sugestoes(
                movimento_banco=movimento_banco,
                movimentos_sistema=movimentos_sistema,
                ids_sistema_usados=ids_sistema_usados,
                tolerancia_valor=tolerancia_valor,
                tolerancia_dias=tolerancia_dias,
            )

            item_banco["sugestoes"] = sugestoes
            pendentes_banco.append(item_banco)

    pendentes_sistema = []

    for movimento_sistema in movimentos_sistema:
        chave_sistema = f"{movimento_sistema['origem']}:{movimento_sistema['id']}"

        if chave_sistema not in ids_sistema_usados:
            pendentes_sistema.append(_serializar_movimento_sistema(movimento_sistema))

    total_banco_entradas = sum(
        (_decimal(item["valor"]) for item in movimentos_banco if item["tipo"] == "ENTRADA"),
        Decimal("0.00"),
    )
    total_banco_saidas = sum(
        (_decimal(item["valor"]) for item in movimentos_banco if item["tipo"] == "SAIDA"),
        Decimal("0.00"),
    )
    total_sistema_entradas = sum(
        (_decimal(item["valor"]) for item in movimentos_sistema if item["tipo"] == "ENTRADA"),
        Decimal("0.00"),
    )
    total_sistema_saidas = sum(
        (_decimal(item["valor"]) for item in movimentos_sistema if item["tipo"] == "SAIDA"),
        Decimal("0.00"),
    )

    return {
        "resumo": {
            "movimentos_banco": len(movimentos_banco),
            "movimentos_sistema": len(movimentos_sistema),
            "conciliados": len(conciliados),
            "pendentes_banco": len(pendentes_banco),
            "pendentes_sistema": len(pendentes_sistema),
            "total_banco_entradas": _float(total_banco_entradas),
            "total_banco_saidas": _float(total_banco_saidas),
            "total_sistema_entradas": _float(total_sistema_entradas),
            "total_sistema_saidas": _float(total_sistema_saidas),
            "diferenca_entradas": _float(total_banco_entradas - total_sistema_entradas),
            "diferenca_saidas": _float(total_banco_saidas - total_sistema_saidas),
        },
        "conciliados": conciliados,
        "pendentes_banco": pendentes_banco,
        "pendentes_sistema": pendentes_sistema,
    }


def _buscar_sugestoes(
    movimento_banco: dict[str, Any],
    movimentos_sistema: list[dict[str, Any]],
    ids_sistema_usados: set[str],
    tolerancia_valor: Decimal,
    tolerancia_dias: int,
) -> list[dict[str, Any]]:
    sugestoes = []

    for movimento_sistema in movimentos_sistema:
        chave_sistema = f"{movimento_sistema['origem']}:{movimento_sistema['id']}"

        if chave_sistema in ids_sistema_usados:
            continue

        score, motivos = _pontuar_match(
            movimento_banco=movimento_banco,
            movimento_sistema=movimento_sistema,
            tolerancia_valor=tolerancia_valor * Decimal("3"),
            tolerancia_dias=tolerancia_dias + 3,
        )

        if score >= 50:
            sugestoes.append(
                {
                    "sistema": _serializar_movimento_sistema(movimento_sistema),
                    "score": score,
                    "motivo": ", ".join(motivos),
                }
            )

    sugestoes.sort(key=lambda item: item["score"], reverse=True)
    return sugestoes[:3]


def _serializar_movimento_banco(item: dict[str, Any]) -> dict[str, Any]:
    return {
        "linha": item.get("linha"),
        "data": item.get("data_iso"),
        "tipo": item.get("tipo"),
        "descricao": item.get("descricao"),
        "documento": item.get("documento"),
        "valor": _float(item.get("valor")),
        "valor_original": _float(item.get("valor_original")),
    }


def _serializar_movimento_sistema(item: dict[str, Any]) -> dict[str, Any]:
    return {
        "origem": item.get("origem"),
        "id": item.get("id"),
        "tipo": item.get("tipo"),
        "data": item.get("data_iso"),
        "descricao": item.get("descricao"),
        "pessoa": item.get("pessoa"),
        "forma_pagamento": item.get("forma_pagamento"),
        "valor": _float(item.get("valor")),
        "status": item.get("status"),
    }


def _periodo_do_arquivo(movimentos: list[dict[str, Any]]) -> tuple[date, date]:
    datas = [item["data"] for item in movimentos if item.get("data")]

    if not datas:
        hoje = date.today()
        return hoje, hoje

    return min(datas), max(datas)


def _gerar_modelo_csv() -> BytesIO:
    linhas = [
        ["Data", "Descrição", "Valor", "Documento"],
        [date.today().strftime("%d/%m/%Y"), "Exemplo de entrada PIX", "150,00", "PIX123"],
        [date.today().strftime("%d/%m/%Y"), "Exemplo de saída fornecedor", "-80,00", "PAG456"],
    ]

    buffer = StringIO()
    writer = csv.writer(buffer, delimiter=";", lineterminator="\n")
    writer.writerows(linhas)

    conteudo = "\ufeff" + buffer.getvalue()
    output = BytesIO(conteudo.encode("utf-8"))
    output.seek(0)
    return output


def _processar_conciliacao(
    db: Session,
    empresa_id: int,
    movimentos_banco: list[dict[str, Any]],
    data_inicio: date | None,
    data_fim: date | None,
    tolerancia_centavos: int,
    tolerancia_dias: int,
) -> dict[str, Any]:
    periodo_inicio_arquivo, periodo_fim_arquivo = _periodo_do_arquivo(movimentos_banco)

    inicio = data_inicio or periodo_inicio_arquivo
    fim = data_fim or periodo_fim_arquivo

    margem_inicio = inicio - timedelta(days=max(tolerancia_dias, 0))
    margem_fim = fim + timedelta(days=max(tolerancia_dias, 0))

    movimentos_sistema = _consultar_financeiro_realizado(
        db=db,
        empresa_id=empresa_id,
        data_inicio=margem_inicio,
        data_fim=margem_fim,
    )

    tolerancia_valor = Decimal(tolerancia_centavos) / Decimal("100")

    resultado = _conciliar_movimentos(
        movimentos_banco=movimentos_banco,
        movimentos_sistema=movimentos_sistema,
        tolerancia_valor=tolerancia_valor,
        tolerancia_dias=tolerancia_dias,
    )

    resultado["empresa_id"] = empresa_id
    resultado["data_inicio"] = inicio.isoformat()
    resultado["data_fim"] = fim.isoformat()
    resultado["periodo_arquivo"] = {
        "data_inicio": periodo_inicio_arquivo.isoformat(),
        "data_fim": periodo_fim_arquivo.isoformat(),
    }
    resultado["parametros"] = {
        "tolerancia_centavos": tolerancia_centavos,
        "tolerancia_valor": _float(tolerancia_valor),
        "tolerancia_dias": tolerancia_dias,
    }

    return resultado


def _salvar_historico_conciliacao(
    db: Session,
    empresa_id: int,
    nome_arquivo: str,
    tipo_arquivo: str,
    resultado: dict[str, Any],
) -> ConciliacaoBancaria:
    resumo = resultado.get("resumo", {})
    parametros = resultado.get("parametros", {})

    registro = ConciliacaoBancaria(
        empresa_id=empresa_id,
        nome_arquivo=nome_arquivo or "extrato_bancario",
        tipo_arquivo=tipo_arquivo.upper(),
        data_inicio=_parse_data(resultado.get("data_inicio")) or date.today(),
        data_fim=_parse_data(resultado.get("data_fim")) or date.today(),
        movimentos_banco=int(resumo.get("movimentos_banco") or 0),
        movimentos_sistema=int(resumo.get("movimentos_sistema") or 0),
        conciliados=int(resumo.get("conciliados") or 0),
        pendentes_banco=int(resumo.get("pendentes_banco") or 0),
        pendentes_sistema=int(resumo.get("pendentes_sistema") or 0),
        total_banco_entradas=_decimal(resumo.get("total_banco_entradas")),
        total_banco_saidas=_decimal(resumo.get("total_banco_saidas")),
        total_sistema_entradas=_decimal(resumo.get("total_sistema_entradas")),
        total_sistema_saidas=_decimal(resumo.get("total_sistema_saidas")),
        diferenca_entradas=_decimal(resumo.get("diferenca_entradas")),
        diferenca_saidas=_decimal(resumo.get("diferenca_saidas")),
        tolerancia_centavos=int(parametros.get("tolerancia_centavos") or 2),
        tolerancia_dias=int(parametros.get("tolerancia_dias") or 2),
        resultado_json=resultado,
    )

    db.add(registro)
    db.commit()
    db.refresh(registro)

    resultado["historico_id"] = registro.id

    return registro


def _serializar_historico_conciliacao(
    registro: ConciliacaoBancaria,
    incluir_resultado: bool = False,
) -> dict[str, Any]:
    dados = {
        "id": registro.id,
        "empresa_id": registro.empresa_id,
        "nome_arquivo": registro.nome_arquivo,
        "tipo_arquivo": registro.tipo_arquivo,
        "data_inicio": registro.data_inicio.isoformat() if registro.data_inicio else None,
        "data_fim": registro.data_fim.isoformat() if registro.data_fim else None,
        "movimentos_banco": registro.movimentos_banco,
        "movimentos_sistema": registro.movimentos_sistema,
        "conciliados": registro.conciliados,
        "pendentes_banco": registro.pendentes_banco,
        "pendentes_sistema": registro.pendentes_sistema,
        "total_banco_entradas": _float(registro.total_banco_entradas),
        "total_banco_saidas": _float(registro.total_banco_saidas),
        "total_sistema_entradas": _float(registro.total_sistema_entradas),
        "total_sistema_saidas": _float(registro.total_sistema_saidas),
        "diferenca_entradas": _float(registro.diferenca_entradas),
        "diferenca_saidas": _float(registro.diferenca_saidas),
        "tolerancia_centavos": registro.tolerancia_centavos,
        "tolerancia_dias": registro.tolerancia_dias,
        "criado_em": registro.criado_em.isoformat() if registro.criado_em else None,
        "atualizado_em": registro.atualizado_em.isoformat() if registro.atualizado_em else None,
    }

    if incluir_resultado:
        dados["resultado"] = registro.resultado_json

    return dados


@router.get("/modelo-csv")
def baixar_modelo_csv():
    arquivo = _gerar_modelo_csv()

    return StreamingResponse(
        arquivo,
        media_type="text/csv; charset=utf-8",
        headers={
            "Content-Disposition": 'attachment; filename="modelo_conciliacao_bancaria.csv"'
        },
    )


@router.post("/importar-csv")
def importar_csv_conciliacao_bancaria(
    empresa_id: int = Query(..., ge=1),
    data_inicio: date | None = Query(None),
    data_fim: date | None = Query(None),
    tolerancia_centavos: int = Query(2, ge=0, le=10000),
    tolerancia_dias: int = Query(2, ge=0, le=30),
    arquivo: UploadFile = File(...),
    db: Session = Depends(get_db),
):
    nome_arquivo = arquivo.filename or ""

    if not nome_arquivo.lower().endswith(".csv"):
        raise HTTPException(
            status_code=400,
            detail="Envie um arquivo CSV.",
        )

    texto = _ler_upload_texto(arquivo)
    movimentos_banco = _parse_csv(texto)

    resultado = _processar_conciliacao(
        db=db,
        empresa_id=empresa_id,
        movimentos_banco=movimentos_banco,
        data_inicio=data_inicio,
        data_fim=data_fim,
        tolerancia_centavos=tolerancia_centavos,
        tolerancia_dias=tolerancia_dias,
    )

    resultado["nome_arquivo"] = nome_arquivo
    resultado["tipo_arquivo"] = "CSV"
    resultado["gravado"] = False

    return resultado

@router.post("/importar-ofx")
def importar_ofx_conciliacao_bancaria(
    empresa_id: int = Query(..., ge=1),
    data_inicio: date | None = Query(None),
    data_fim: date | None = Query(None),
    tolerancia_centavos: int = Query(2, ge=0, le=10000),
    tolerancia_dias: int = Query(2, ge=0, le=30),
    arquivo: UploadFile = File(...),
    db: Session = Depends(get_db),
):
    nome_arquivo = arquivo.filename or ""

    if not nome_arquivo.lower().endswith((".ofx", ".qfx")):
        raise HTTPException(
            status_code=400,
            detail="Envie um arquivo OFX ou QFX.",
        )

    texto = _ler_upload_texto(arquivo)
    movimentos_banco = _parse_ofx(texto)

    resultado = _processar_conciliacao(
        db=db,
        empresa_id=empresa_id,
        movimentos_banco=movimentos_banco,
        data_inicio=data_inicio,
        data_fim=data_fim,
        tolerancia_centavos=tolerancia_centavos,
        tolerancia_dias=tolerancia_dias,
    )

    resultado["nome_arquivo"] = nome_arquivo
    resultado["tipo_arquivo"] = "OFX"
    resultado["gravado"] = False

    return resultado

@router.post("/cadastrar")
def cadastrar_conciliacao_bancaria(
    payload: dict[str, Any] = Body(...),
    db: Session = Depends(get_db),
):
    empresa_id = int(payload.get("empresa_id") or 0)
    resultado = payload.get("resultado") or {}
    nome_arquivo = payload.get("nome_arquivo") or resultado.get("nome_arquivo") or "extrato_bancario"
    tipo_arquivo = payload.get("tipo_arquivo") or resultado.get("tipo_arquivo") or "CSV"

    if empresa_id <= 0:
        raise HTTPException(
            status_code=400,
            detail="Empresa inválida para cadastrar a conciliação.",
        )

    if not resultado or not isinstance(resultado, dict):
        raise HTTPException(
            status_code=400,
            detail="Resultado da conciliação não informado.",
        )

    resumo = resultado.get("resumo") or {}

    if not resumo:
        raise HTTPException(
            status_code=400,
            detail="Resumo da conciliação não informado.",
        )

    if resultado.get("historico_id"):
        existente = (
            db.query(ConciliacaoBancaria)
            .filter(ConciliacaoBancaria.id == int(resultado["historico_id"]))
            .filter(ConciliacaoBancaria.empresa_id == empresa_id)
            .first()
        )

        if existente:
            return {
                "ok": True,
                "ja_cadastrado": True,
                "historico": _serializar_historico_conciliacao(existente),
            }

    registro = _salvar_historico_conciliacao(
        db=db,
        empresa_id=empresa_id,
        nome_arquivo=nome_arquivo,
        tipo_arquivo=tipo_arquivo,
        resultado=resultado,
    )

    return {
        "ok": True,
        "ja_cadastrado": False,
        "historico": _serializar_historico_conciliacao(registro),
    }


@router.get("/financeiro-realizado")
def listar_financeiro_realizado_para_conciliacao(
    empresa_id: int = Query(..., ge=1),
    data_inicio: date | None = Query(None),
    data_fim: date | None = Query(None),
    db: Session = Depends(get_db),
):
    hoje = date.today()
    inicio = data_inicio or hoje.replace(day=1)
    fim = data_fim or hoje

    if fim < inicio:
        raise HTTPException(
            status_code=400,
            detail="A data final não pode ser menor que a data inicial.",
        )

    movimentos = _consultar_financeiro_realizado(
        db=db,
        empresa_id=empresa_id,
        data_inicio=inicio,
        data_fim=fim,
    )

    entradas = sum(
        (_decimal(item["valor"]) for item in movimentos if item["tipo"] == "ENTRADA"),
        Decimal("0.00"),
    )
    saidas = sum(
        (_decimal(item["valor"]) for item in movimentos if item["tipo"] == "SAIDA"),
        Decimal("0.00"),
    )

    return {
        "empresa_id": empresa_id,
        "data_inicio": inicio.isoformat(),
        "data_fim": fim.isoformat(),
        "resumo": {
            "movimentos": len(movimentos),
            "entradas": _float(entradas),
            "saidas": _float(saidas),
            "saldo": _float(entradas - saidas),
        },
        "movimentos": [_serializar_movimento_sistema(item) for item in movimentos],
    }

def _formatar_data_excel(valor: Any) -> str:
    if not valor:
        return ""

    if isinstance(valor, date):
        return valor.strftime("%d/%m/%Y")

    texto = str(valor)

    try:
        partes = texto.split("-")
        if len(partes) == 3:
            return f"{partes[2]}/{partes[1]}/{partes[0]}"
    except Exception:
        pass

    return texto


def _xlsx_aplicar_cabecalho(ws, linha: int, coluna_inicial: int, coluna_final: int) -> None:
    fill = PatternFill("solid", fgColor="0F172A")
    font = Font(color="FFFFFF", bold=True)
    alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    border = Border(bottom=Side(style="thin", color="CBD5E1"))

    for coluna in range(coluna_inicial, coluna_final + 1):
        cell = ws.cell(row=linha, column=coluna)
        cell.fill = fill
        cell.font = font
        cell.alignment = alignment
        cell.border = border


def _xlsx_aplicar_bordas(ws, linha_inicial: int, linha_final: int, coluna_inicial: int, coluna_final: int) -> None:
    border = Border(
        left=Side(style="thin", color="E2E8F0"),
        right=Side(style="thin", color="E2E8F0"),
        top=Side(style="thin", color="E2E8F0"),
        bottom=Side(style="thin", color="E2E8F0"),
    )

    for linha in range(linha_inicial, linha_final + 1):
        for coluna in range(coluna_inicial, coluna_final + 1):
            ws.cell(row=linha, column=coluna).border = border


def _xlsx_ajustar_larguras(ws) -> None:
    for coluna_cells in ws.columns:
        max_length = 0
        coluna_letra = get_column_letter(coluna_cells[0].column)

        for cell in coluna_cells:
            if cell.value is None:
                continue

            max_length = max(max_length, len(str(cell.value)))

        ws.column_dimensions[coluna_letra].width = min(max(max_length + 3, 12), 46)


def _xlsx_formatar_moeda(ws, colunas: list[int], linha_inicial: int, linha_final: int) -> None:
    for linha in range(linha_inicial, linha_final + 1):
        for coluna in colunas:
            ws.cell(row=linha, column=coluna).number_format = 'R$ #,##0.00;[Red]-R$ #,##0.00'


def _xlsx_estilizar_aba(ws) -> None:
    for linha in ws.iter_rows():
        for cell in linha:
            cell.alignment = Alignment(
                vertical="center",
                horizontal=cell.alignment.horizontal or "left",
                wrap_text=True,
            )

    _xlsx_ajustar_larguras(ws)


def _criar_xlsx_conciliacao_bancaria(resultado: dict[str, Any]) -> BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "Resumo"

    resumo = resultado.get("resumo", {})
    parametros = resultado.get("parametros", {})
    periodo_arquivo = resultado.get("periodo_arquivo", {})

    ws["A1"] = "Conciliação Bancária"
    ws["A1"].font = Font(size=18, bold=True, color="0F172A")
    ws["A2"] = f"Empresa: {resultado.get('empresa_id')}"
    ws["A3"] = f"Período considerado: {_formatar_data_excel(resultado.get('data_inicio'))} a {_formatar_data_excel(resultado.get('data_fim'))}"
    ws["A4"] = f"Período do arquivo: {_formatar_data_excel(periodo_arquivo.get('data_inicio'))} a {_formatar_data_excel(periodo_arquivo.get('data_fim'))}"
    ws["A5"] = f"Tolerância: {parametros.get('tolerancia_dias', 0)} dia(s) e R$ {parametros.get('tolerancia_valor', 0)}"
    ws["A6"] = f"Gerado em: {datetime.now().strftime('%d/%m/%Y %H:%M')}"

    linha_resumo = 8
    ws.cell(row=linha_resumo, column=1, value="Indicador")
    ws.cell(row=linha_resumo, column=2, value="Valor")
    _xlsx_aplicar_cabecalho(ws, linha_resumo, 1, 2)

    linhas_resumo = [
        ("Movimentos no banco", resumo.get("movimentos_banco", 0)),
        ("Movimentos no sistema", resumo.get("movimentos_sistema", 0)),
        ("Conciliados", resumo.get("conciliados", 0)),
        ("Pendentes no banco", resumo.get("pendentes_banco", 0)),
        ("Pendentes no sistema", resumo.get("pendentes_sistema", 0)),
        ("Total banco entradas", resumo.get("total_banco_entradas", 0)),
        ("Total banco saídas", resumo.get("total_banco_saidas", 0)),
        ("Total sistema entradas", resumo.get("total_sistema_entradas", 0)),
        ("Total sistema saídas", resumo.get("total_sistema_saidas", 0)),
        ("Diferença entradas", resumo.get("diferenca_entradas", 0)),
        ("Diferença saídas", resumo.get("diferenca_saidas", 0)),
    ]

    for index, item in enumerate(linhas_resumo, start=linha_resumo + 1):
        ws.cell(row=index, column=1, value=item[0])
        ws.cell(row=index, column=2, value=item[1])

    _xlsx_formatar_moeda(ws, [2], linha_resumo + 6, linha_resumo + len(linhas_resumo))
    _xlsx_aplicar_bordas(ws, linha_resumo, linha_resumo + len(linhas_resumo), 1, 2)

    ws_conciliados = wb.create_sheet("Conciliados")
    ws_conciliados["A1"] = "Movimentos conciliados"
    ws_conciliados["A1"].font = Font(size=16, bold=True, color="0F172A")

    headers_conciliados = [
        "Data banco",
        "Descrição banco",
        "Documento banco",
        "Tipo banco",
        "Valor banco",
        "Data sistema",
        "Descrição sistema",
        "Pessoa",
        "Tipo sistema",
        "Forma de pagamento",
        "Valor sistema",
        "Score",
        "Motivo",
    ]

    for col, titulo in enumerate(headers_conciliados, start=1):
        ws_conciliados.cell(row=3, column=col, value=titulo)

    _xlsx_aplicar_cabecalho(ws_conciliados, 3, 1, len(headers_conciliados))

    conciliados = resultado.get("conciliados", [])

    for row_index, item in enumerate(conciliados, start=4):
        banco = item.get("banco", {})
        sistema = item.get("sistema", {})

        ws_conciliados.cell(row=row_index, column=1, value=_formatar_data_excel(banco.get("data")))
        ws_conciliados.cell(row=row_index, column=2, value=banco.get("descricao"))
        ws_conciliados.cell(row=row_index, column=3, value=banco.get("documento"))
        ws_conciliados.cell(row=row_index, column=4, value=banco.get("tipo"))
        ws_conciliados.cell(row=row_index, column=5, value=banco.get("valor"))
        ws_conciliados.cell(row=row_index, column=6, value=_formatar_data_excel(sistema.get("data")))
        ws_conciliados.cell(row=row_index, column=7, value=sistema.get("descricao"))
        ws_conciliados.cell(row=row_index, column=8, value=sistema.get("pessoa"))
        ws_conciliados.cell(row=row_index, column=9, value=sistema.get("tipo"))
        ws_conciliados.cell(row=row_index, column=10, value=sistema.get("forma_pagamento"))
        ws_conciliados.cell(row=row_index, column=11, value=sistema.get("valor"))
        ws_conciliados.cell(row=row_index, column=12, value=item.get("score"))
        ws_conciliados.cell(row=row_index, column=13, value=item.get("motivo"))

    ultima_linha_conciliados = max(4, 3 + len(conciliados))
    _xlsx_formatar_moeda(ws_conciliados, [5, 11], 4, ultima_linha_conciliados)
    _xlsx_aplicar_bordas(ws_conciliados, 3, ultima_linha_conciliados, 1, len(headers_conciliados))
    ws_conciliados.freeze_panes = "A4"
    ws_conciliados.auto_filter.ref = f"A3:M{ultima_linha_conciliados}"

    ws_pendentes_banco = wb.create_sheet("Pend Banco")
    ws_pendentes_banco["A1"] = "Pendentes no banco"
    ws_pendentes_banco["A1"].font = Font(size=16, bold=True, color="0F172A")

    headers_pendentes_banco = [
        "Data",
        "Descrição",
        "Documento",
        "Tipo",
        "Valor",
        "Qtd. sugestões",
        "Sugestão 1",
        "Score 1",
        "Motivo 1",
        "Sugestão 2",
        "Score 2",
        "Motivo 2",
        "Sugestão 3",
        "Score 3",
        "Motivo 3",
    ]

    for col, titulo in enumerate(headers_pendentes_banco, start=1):
        ws_pendentes_banco.cell(row=3, column=col, value=titulo)

    _xlsx_aplicar_cabecalho(ws_pendentes_banco, 3, 1, len(headers_pendentes_banco))

    pendentes_banco = resultado.get("pendentes_banco", [])

    for row_index, item in enumerate(pendentes_banco, start=4):
        sugestoes = item.get("sugestoes", [])

        ws_pendentes_banco.cell(row=row_index, column=1, value=_formatar_data_excel(item.get("data")))
        ws_pendentes_banco.cell(row=row_index, column=2, value=item.get("descricao"))
        ws_pendentes_banco.cell(row=row_index, column=3, value=item.get("documento"))
        ws_pendentes_banco.cell(row=row_index, column=4, value=item.get("tipo"))
        ws_pendentes_banco.cell(row=row_index, column=5, value=item.get("valor"))
        ws_pendentes_banco.cell(row=row_index, column=6, value=len(sugestoes))

        coluna = 7
        for sugestao in sugestoes[:3]:
            sistema = sugestao.get("sistema", {})
            descricao_sugestao = f"{_formatar_data_excel(sistema.get('data'))} | {sistema.get('descricao') or ''} | R$ {sistema.get('valor', 0)}"

            ws_pendentes_banco.cell(row=row_index, column=coluna, value=descricao_sugestao)
            ws_pendentes_banco.cell(row=row_index, column=coluna + 1, value=sugestao.get("score"))
            ws_pendentes_banco.cell(row=row_index, column=coluna + 2, value=sugestao.get("motivo"))

            coluna += 3

    ultima_linha_pend_banco = max(4, 3 + len(pendentes_banco))
    _xlsx_formatar_moeda(ws_pendentes_banco, [5], 4, ultima_linha_pend_banco)
    _xlsx_aplicar_bordas(ws_pendentes_banco, 3, ultima_linha_pend_banco, 1, len(headers_pendentes_banco))
    ws_pendentes_banco.freeze_panes = "A4"
    ws_pendentes_banco.auto_filter.ref = f"A3:O{ultima_linha_pend_banco}"

    ws_pendentes_sistema = wb.create_sheet("Pend Sistema")
    ws_pendentes_sistema["A1"] = "Pendentes no sistema"
    ws_pendentes_sistema["A1"].font = Font(size=16, bold=True, color="0F172A")

    headers_pendentes_sistema = [
        "Data",
        "Descrição",
        "Pessoa",
        "Tipo",
        "Forma de pagamento",
        "Valor",
        "Status",
        "Origem",
        "ID",
    ]

    for col, titulo in enumerate(headers_pendentes_sistema, start=1):
        ws_pendentes_sistema.cell(row=3, column=col, value=titulo)

    _xlsx_aplicar_cabecalho(ws_pendentes_sistema, 3, 1, len(headers_pendentes_sistema))

    pendentes_sistema = resultado.get("pendentes_sistema", [])

    for row_index, item in enumerate(pendentes_sistema, start=4):
        ws_pendentes_sistema.cell(row=row_index, column=1, value=_formatar_data_excel(item.get("data")))
        ws_pendentes_sistema.cell(row=row_index, column=2, value=item.get("descricao"))
        ws_pendentes_sistema.cell(row=row_index, column=3, value=item.get("pessoa"))
        ws_pendentes_sistema.cell(row=row_index, column=4, value=item.get("tipo"))
        ws_pendentes_sistema.cell(row=row_index, column=5, value=item.get("forma_pagamento"))
        ws_pendentes_sistema.cell(row=row_index, column=6, value=item.get("valor"))
        ws_pendentes_sistema.cell(row=row_index, column=7, value=item.get("status"))
        ws_pendentes_sistema.cell(row=row_index, column=8, value=item.get("origem"))
        ws_pendentes_sistema.cell(row=row_index, column=9, value=item.get("id"))

    ultima_linha_pend_sistema = max(4, 3 + len(pendentes_sistema))
    _xlsx_formatar_moeda(ws_pendentes_sistema, [6], 4, ultima_linha_pend_sistema)
    _xlsx_aplicar_bordas(ws_pendentes_sistema, 3, ultima_linha_pend_sistema, 1, len(headers_pendentes_sistema))
    ws_pendentes_sistema.freeze_panes = "A4"
    ws_pendentes_sistema.auto_filter.ref = f"A3:I{ultima_linha_pend_sistema}"

    for sheet in wb.worksheets:
        _xlsx_estilizar_aba(sheet)

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


@router.post("/exportar-xlsx")
def exportar_xlsx_conciliacao_bancaria(
    empresa_id: int = Query(..., ge=1),
    data_inicio: date | None = Query(None),
    data_fim: date | None = Query(None),
    tolerancia_centavos: int = Query(2, ge=0, le=10000),
    tolerancia_dias: int = Query(2, ge=0, le=30),
    arquivo: UploadFile = File(...),
    db: Session = Depends(get_db),
):
    nome_arquivo = arquivo.filename or ""
    nome_lower = nome_arquivo.lower()

    if nome_lower.endswith(".csv"):
        texto = _ler_upload_texto(arquivo)
        movimentos_banco = _parse_csv(texto)
    elif nome_lower.endswith((".ofx", ".qfx")):
        texto = _ler_upload_texto(arquivo)
        movimentos_banco = _parse_ofx(texto)
    else:
        raise HTTPException(
            status_code=400,
            detail="Envie um arquivo CSV, OFX ou QFX.",
        )

    resultado = _processar_conciliacao(
        db=db,
        empresa_id=empresa_id,
        movimentos_banco=movimentos_banco,
        data_inicio=data_inicio,
        data_fim=data_fim,
        tolerancia_centavos=tolerancia_centavos,
        tolerancia_dias=tolerancia_dias,
    )

    arquivo_xlsx = _criar_xlsx_conciliacao_bancaria(resultado)
    nome_saida = f"conciliacao_bancaria_empresa_{empresa_id}_{resultado['data_inicio']}_{resultado['data_fim']}.xlsx"

    return StreamingResponse(
        arquivo_xlsx,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f'attachment; filename="{nome_saida}"'
        },
    )


@router.get("/historico")
def listar_historico_conciliacoes_bancarias(
    empresa_id: int = Query(..., ge=1),
    data_inicio: date | None = Query(None),
    data_fim: date | None = Query(None),
    limite: int = Query(30, ge=1, le=200),
    db: Session = Depends(get_db),
):
    query = (
        db.query(ConciliacaoBancaria)
        .filter(ConciliacaoBancaria.empresa_id == empresa_id)
    )

    if data_inicio:
        query = query.filter(ConciliacaoBancaria.data_fim >= data_inicio)

    if data_fim:
        query = query.filter(ConciliacaoBancaria.data_inicio <= data_fim)

    registros = (
        query
        .order_by(ConciliacaoBancaria.criado_em.desc(), ConciliacaoBancaria.id.desc())
        .limit(limite)
        .all()
    )

    return {
        "empresa_id": empresa_id,
        "total": len(registros),
        "historico": [
            _serializar_historico_conciliacao(registro)
            for registro in registros
        ],
    }


@router.get("/historico/{conciliacao_id}")
def obter_historico_conciliacao_bancaria(
    conciliacao_id: int,
    empresa_id: int = Query(..., ge=1),
    db: Session = Depends(get_db),
):
    registro = (
        db.query(ConciliacaoBancaria)
        .filter(ConciliacaoBancaria.id == conciliacao_id)
        .filter(ConciliacaoBancaria.empresa_id == empresa_id)
        .first()
    )

    if not registro:
        raise HTTPException(
            status_code=404,
            detail="Conciliação bancária não encontrada.",
        )

    return _serializar_historico_conciliacao(registro, incluir_resultado=True)


@router.get("/historico/{conciliacao_id}/exportar-xlsx")
def exportar_xlsx_historico_conciliacao_bancaria(
    conciliacao_id: int,
    empresa_id: int = Query(..., ge=1),
    db: Session = Depends(get_db),
):
    registro = (
        db.query(ConciliacaoBancaria)
        .filter(ConciliacaoBancaria.id == conciliacao_id)
        .filter(ConciliacaoBancaria.empresa_id == empresa_id)
        .first()
    )

    if not registro:
        raise HTTPException(
            status_code=404,
            detail="Conciliação bancária não encontrada.",
        )

    resultado = registro.resultado_json
    arquivo_xlsx = _criar_xlsx_conciliacao_bancaria(resultado)
    nome_saida = f"conciliacao_bancaria_historico_{registro.id}_empresa_{empresa_id}.xlsx"

    return StreamingResponse(
        arquivo_xlsx,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f'attachment; filename="{nome_saida}"'
        },
    )



def _conciliacao_tabela_existe(db: Session, nome_tabela: str) -> bool:
    resultado = db.execute(
        text("select to_regclass(:nome_tabela)"),
        {"nome_tabela": nome_tabela},
    ).scalar()

    return resultado is not None


def _conciliacao_colunas_tabela(db: Session, nome_tabela: str) -> set[str]:
    if not _conciliacao_tabela_existe(db, nome_tabela):
        return set()

    linhas = db.execute(
        text("""
            select column_name
            from information_schema.columns
            where table_name = :nome_tabela
        """),
        {"nome_tabela": nome_tabela},
    ).all()

    return {linha[0] for linha in linhas}


@router.get("/dre-opcoes")
def listar_dre_opcoes_conciliacao_bancaria(
    empresa_id: int = Query(..., ge=1),
    db: Session = Depends(get_db),
):
    if not _conciliacao_tabela_existe(db, "financeiro_plano_dre"):
        return {
            "empresa_id": empresa_id,
            "opcoes": [],
        }

    colunas = _conciliacao_colunas_tabela(db, "financeiro_plano_dre")

    campos = ["id"]

    for campo in ["grupo", "categoria", "subcategoria", "nome", "descricao"]:
        if campo in colunas:
            campos.append(campo)

    where = []
    params = {"empresa_id": empresa_id}

    if "empresa_id" in colunas:
        where.append("(empresa_id = :empresa_id or empresa_id is null)")

    if "ativo" in colunas:
        where.append("(ativo = true or ativo is null)")

    where_sql = " where " + " and ".join(where) if where else ""

    ordem = [
        campo
        for campo in ["grupo", "categoria", "subcategoria", "nome", "id"]
        if campo in campos
    ]

    order_sql = " order by " + (", ".join(ordem) if ordem else "id")

    sql = f"select {', '.join(campos)} from financeiro_plano_dre {where_sql} {order_sql}"

    rows = db.execute(text(sql), params).mappings().all()

    opcoes = []

    for row in rows:
        partes = []

        for campo in ["grupo", "categoria", "subcategoria", "nome", "descricao"]:
            valor = row.get(campo)

            if valor and str(valor).strip():
                texto = str(valor).strip()

                if texto not in partes:
                    partes.append(texto)

        label = " > ".join(partes) if partes else f"Classificação #{row['id']}"

        opcoes.append(
            {
                "id": row["id"],
                "label": label,
                "grupo": row.get("grupo"),
                "categoria": row.get("categoria"),
                "subcategoria": row.get("subcategoria"),
            }
        )

    return {
        "empresa_id": empresa_id,
        "total": len(opcoes),
        "opcoes": opcoes,
    }



def _conciliacao_primeiro_id_tabela(db: Session, nome_tabela: str, empresa_id: int | None = None) -> int | None:
    if not _conciliacao_tabela_existe(db, nome_tabela):
        return None

    colunas = _conciliacao_colunas_tabela(db, nome_tabela)

    if empresa_id and "empresa_id" in colunas:
        row = db.execute(
            text(f"select id from {nome_tabela} where empresa_id = :empresa_id order by id limit 1"),
            {"empresa_id": empresa_id},
        ).first()

        if row:
            return int(row[0])

    row = db.execute(text(f"select id from {nome_tabela} order by id limit 1")).first()
    return int(row[0]) if row else None


def _conciliacao_colunas_modelo(modelo: Any) -> dict[str, Any]:
    return {coluna.name: coluna for coluna in modelo.__table__.columns}


def _conciliacao_preencher_obrigatorios(
    db: Session,
    modelo: Any,
    dados: dict[str, Any],
    empresa_id: int,
) -> dict[str, Any]:
    colunas = _conciliacao_colunas_modelo(modelo)

    for nome, coluna in colunas.items():
        if nome in dados:
            continue

        if coluna.primary_key:
            continue

        if coluna.nullable:
            continue

        if coluna.default is not None or coluna.server_default is not None:
            continue

        fks = list(coluna.foreign_keys)

        if fks:
            tabela = fks[0].column.table.name
            primeiro_id = _conciliacao_primeiro_id_tabela(db, tabela, empresa_id)

            if primeiro_id:
                dados[nome] = primeiro_id
                continue

        tipo = str(coluna.type).upper()

        if "DATE" in tipo:
            dados[nome] = date.today()
        elif "NUMERIC" in tipo or "DECIMAL" in tipo or "FLOAT" in tipo:
            dados[nome] = Decimal("0.00")
        elif "INTEGER" in tipo:
            dados[nome] = 0
        elif "BOOLEAN" in tipo:
            dados[nome] = False
        else:
            dados[nome] = "CONCILIACAO"

    return dados


def _conciliacao_dados_base_lancamento(
    db: Session,
    modelo: Any,
    empresa_id: int,
    banco: dict[str, Any],
    descricao_prefixo: str,
) -> dict[str, Any]:
    colunas = _conciliacao_colunas_modelo(modelo)

    data_banco = _parse_data(banco.get("data")) or date.today()
    valor = _decimal(banco.get("valor"))
    descricao_banco = str(banco.get("descricao") or "Movimento bancário").strip()
    documento = str(banco.get("documento") or "").strip()

    descricao = f"{descricao_prefixo} - {descricao_banco}"

    if documento and documento != "-":
        descricao = f"{descricao} - Doc. {documento}"

    dados: dict[str, Any] = {}

    if "empresa_id" in colunas:
        dados["empresa_id"] = empresa_id

    if "descricao" in colunas:
        dados["descricao"] = descricao[:255]

    if "valor" in colunas:
        dados["valor"] = valor

    if "valor_pago" in colunas:
        dados["valor_pago"] = valor

    if "status" in colunas:
        dados["status"] = "PAGO"

    if "data_pagamento" in colunas:
        dados["data_pagamento"] = data_banco

    if "vencimento" in colunas:
        dados["vencimento"] = data_banco

    if "data_vencimento" in colunas:
        dados["data_vencimento"] = data_banco

    if "competencia" in colunas:
        dados["competencia"] = data_banco

    if "data_competencia" in colunas:
        dados["data_competencia"] = data_banco

    if "forma_pagamento" in colunas:
        dados["forma_pagamento"] = "CONTA_BANCARIA"

    if "forma_pagamento_baixa" in colunas:
        dados["forma_pagamento_baixa"] = "CONTA_BANCARIA"

    if "observacao" in colunas:
        dados["observacao"] = "Lançamento criado pela conciliação bancária."

    if "observacao_baixa" in colunas:
        dados["observacao_baixa"] = "Baixa criada pela conciliação bancária."

    return _conciliacao_preencher_obrigatorios(db, modelo, dados, empresa_id)


def _conciliacao_serializar_registro_criado(registro: Any, tipo: str) -> dict[str, Any]:
    return {
        "origem": "financeiro_pagar" if tipo == "SAIDA" else "financeiro_receber",
        "id": registro.id,
        "tipo": tipo,
        "data": registro.data_pagamento.isoformat() if getattr(registro, "data_pagamento", None) else None,
        "descricao": getattr(registro, "descricao", None),
        "pessoa": getattr(registro, "fornecedor", None) if tipo == "SAIDA" else _nome_cliente(registro),
        "forma_pagamento": _forma_pagamento(registro),
        "valor": _float(_valor_realizado(registro)),
        "status": _status_atual(registro),
    }


@router.post("/lancar-pendentes")
def lancar_pendentes_banco_no_financeiro(
    payload: dict[str, Any] = Body(...),
    db: Session = Depends(get_db),
):
    empresa_id = int(payload.get("empresa_id") or 0)
    itens = payload.get("itens") or []

    if empresa_id <= 0:
        raise HTTPException(
            status_code=400,
            detail="Empresa inválida para lançar pendentes.",
        )

    if not itens:
        raise HTTPException(
            status_code=400,
            detail="Nenhum item selecionado para lançamento.",
        )

    conciliados = []
    erros = []

    for indice, item in enumerate(itens):
        banco = item.get("banco") or item
        tipo = str(banco.get("tipo") or item.get("tipo") or "").upper()
        valor = _decimal(banco.get("valor"))

        if tipo not in ["ENTRADA", "SAIDA"]:
            erros.append({"indice": indice, "erro": "Tipo inválido.", "banco": banco})
            continue

        if valor <= 0:
            erros.append({"indice": indice, "erro": "Valor inválido.", "banco": banco})
            continue

        try:
            if tipo == "SAIDA":
                classificacao_dre_id = item.get("classificacao_dre_id")

                if not classificacao_dre_id:
                    erros.append(
                        {
                            "indice": indice,
                            "erro": "Classificação DRE obrigatória para saída.",
                            "banco": banco,
                        }
                    )
                    continue

                dados = _conciliacao_dados_base_lancamento(
                    db=db,
                    modelo=FinanceiroPagar,
                    empresa_id=empresa_id,
                    banco=banco,
                    descricao_prefixo="Conciliação bancária",
                )

                colunas = _conciliacao_colunas_modelo(FinanceiroPagar)

                if "fornecedor" in colunas:
                    dados["fornecedor"] = str(item.get("fornecedor") or banco.get("descricao") or "Fornecedor bancário")[:255]

                if "classificacao_dre_id" in colunas:
                    dados["classificacao_dre_id"] = int(classificacao_dre_id)

                registro = FinanceiroPagar(**dados)
                db.add(registro)
                db.flush()

                conciliados.append(
                    {
                        "banco": banco,
                        "sistema": _conciliacao_serializar_registro_criado(registro, "SAIDA"),
                        "score": 100,
                        "motivo": "lançado no financeiro pela conciliação bancária",
                        "manual": True,
                        "lancado_financeiro": True,
                    }
                )

            else:
                dados = _conciliacao_dados_base_lancamento(
                    db=db,
                    modelo=FinanceiroReceber,
                    empresa_id=empresa_id,
                    banco=banco,
                    descricao_prefixo="Conciliação bancária",
                )

                colunas = _conciliacao_colunas_modelo(FinanceiroReceber)
                cliente_id = item.get("cliente_id") or _conciliacao_primeiro_id_tabela(db, "clientes", empresa_id)

                if "cliente_id" in colunas and cliente_id:
                    dados["cliente_id"] = int(cliente_id)

                if "cliente_nome" in colunas:
                    dados["cliente_nome"] = str(item.get("cliente_nome") or "Recebimento bancário")[:255]

                registro = FinanceiroReceber(**dados)
                db.add(registro)
                db.flush()

                conciliados.append(
                    {
                        "banco": banco,
                        "sistema": _conciliacao_serializar_registro_criado(registro, "ENTRADA"),
                        "score": 100,
                        "motivo": "lançado no financeiro pela conciliação bancária",
                        "manual": True,
                        "lancado_financeiro": True,
                    }
                )

        except Exception as exc:
            erros.append({"indice": indice, "erro": str(exc), "banco": banco})

    if erros:
        db.rollback()
        raise HTTPException(
            status_code=400,
            detail={
                "mensagem": "Alguns itens não puderam ser lançados.",
                "erros": erros,
            },
        )

    db.commit()

    return {
        "ok": True,
        "empresa_id": empresa_id,
        "quantidade": len(conciliados),
        "conciliados": conciliados,
    }

