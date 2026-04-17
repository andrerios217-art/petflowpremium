from __future__ import annotations

from datetime import date, datetime
from io import BytesIO
from zoneinfo import ZoneInfo

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from sqlalchemy.orm import Session, joinedload

from app.core.deps import get_db, get_empresa_id_atual
from app.models.agendamento import Agendamento
from app.models.agendamento_servico import AgendamentoServico
from app.models.assinatura_pet_consumo import AssinaturaPetConsumo
from app.models.producao import Producao
from app.models.producao_historico import ProducaoHistorico

router = APIRouter(
    prefix="/api/relatorios/banho-tosa",
    tags=["Relatórios - Banho e Tosa"],
)

TZ_LOCAL = ZoneInfo("America/Sao_Paulo")

COLUNAS_RELATORIO = [
    "Data",
    "Horário de Entrada",
    "Tipo de Serviço",
    "Categoria do Cliente",
    "Tutor",
    "Telefone",
    "Pet",
    "Raça do Pet",
    "Grupo do Pet",
    "Ordem de Serviço",
    "Prioridade",
    "Início da Linha",
    "Pré-Banho",
    "Banho",
    "Finalização do Banho",
    "Pré-Tosa",
    "Tosa",
    "Finalização da Tosa",
    "Horário de Finalização",
    "Intercorrência",
]


def _parse_date_param(valor: str | None, nome_campo: str) -> date:
    if not valor:
        raise HTTPException(status_code=400, detail=f"{nome_campo} é obrigatório. Use YYYY-MM-DD.")
    try:
        return datetime.strptime(valor, "%Y-%m-%d").date()
    except ValueError as exc:
        raise HTTPException(
            status_code=400,
            detail=f"{nome_campo} inválido. Use YYYY-MM-DD.",
        ) from exc


def _to_local_datetime(valor):
    if not valor:
        return None
    if isinstance(valor, datetime):
        if valor.tzinfo is None:
            return valor.replace(tzinfo=TZ_LOCAL)
        return valor.astimezone(TZ_LOCAL)
    return None


def _formatar_data(valor) -> str:
    if not valor:
        return "N/A"
    if isinstance(valor, datetime):
        valor = _to_local_datetime(valor)
        if valor:
            return valor.strftime("%d/%m/%Y")
        return "N/A"
    if hasattr(valor, "strftime"):
        return valor.strftime("%d/%m/%Y")
    return str(valor)


def _formatar_hora(valor) -> str:
    if not valor:
        return "N/A"
    if isinstance(valor, datetime):
        valor = _to_local_datetime(valor)
        if valor:
            return valor.strftime("%H:%M:%S")
        return "N/A"
    if hasattr(valor, "strftime"):
        try:
            return valor.strftime("%H:%M:%S")
        except Exception:
            return str(valor)
    return str(valor)


def _texto_limpo(valor, fallback: str = "N/A") -> str:
    if valor is None:
        return fallback
    texto = str(valor).strip()
    return texto if texto else fallback


def _normalizar_prioridade(valor: str | None) -> str:
    texto = _texto_limpo(valor, "Normal").upper()
    mapa = {
        "NORMAL": "Normal",
        "ALTO": "Alto",
        "ALTA": "Alta",
        "URGENTE": "Urgente",
        "BAIXO": "Baixo",
        "BAIXA": "Baixa",
    }
    return mapa.get(texto, texto.title())


def _categoria_cliente(agendamento: Agendamento) -> str:
    created_at_local = _to_local_datetime(getattr(agendamento, "created_at", None))
    data_atendimento = getattr(agendamento, "data", None)

    if created_at_local and data_atendimento:
        return "Check-In" if created_at_local.date() == data_atendimento else "Agendamento"

    return "Agendamento"


def _tem_consumo_assinatura(agendamento_id: int, consumos_por_agendamento: set[int]) -> bool:
    return agendamento_id in consumos_por_agendamento


def _tipo_servico(agendamento_id: int, consumos_por_agendamento: set[int]) -> str:
    return "Assinatura/Pacote" if _tem_consumo_assinatura(agendamento_id, consumos_por_agendamento) else "Avulso"


def _ordem_servico(agendamento: Agendamento) -> str:
    nomes = []
    vistos = set()

    for item in getattr(agendamento, "servicos_agendamento", []) or []:
        servico = getattr(item, "servico", None)
        nome = getattr(servico, "nome", None)
        if nome and nome not in vistos:
            nomes.append(nome)
            vistos.add(nome)

    return ", ".join(nomes) if nomes else "N/A"


def _historicos_ordenados(producao: Producao | None):
    if not producao:
        return []

    historicos = list(getattr(producao, "historicos", []) or [])
    historicos.sort(
        key=lambda h: (
            getattr(h, "iniciado_em", None) is None,
            getattr(h, "iniciado_em", None),
            getattr(h, "id", 0),
        )
    )
    return historicos


def _mapear_historicos_por_etapa(producao: Producao | None) -> dict[str, list]:
    etapas: dict[str, list] = {}
    for historico in _historicos_ordenados(producao):
        etapa = _texto_limpo(getattr(historico, "etapa", None), "")
        if not etapa:
            continue
        etapas.setdefault(etapa, []).append(historico)
    return etapas


def _nome_funcionario_do_historico(historico) -> str:
    funcionario = getattr(historico, "funcionario", None)
    return _texto_limpo(getattr(funcionario, "nome", None), "N/A")


def _nome_primeiro_responsavel(etapas: dict[str, list], etapa: str) -> str:
    itens = etapas.get(etapa, [])
    for historico in itens:
        nome = _nome_funcionario_do_historico(historico)
        if nome != "N/A":
            return nome
    return "N/A"


def _inicio_linha(producao: Producao | None) -> str:
    historicos = _historicos_ordenados(producao)
    for historico in historicos:
        iniciado_em = getattr(historico, "iniciado_em", None)
        if iniciado_em:
            return _formatar_hora(iniciado_em)
    return "N/A"


def _horario_finalizacao(producao: Producao | None) -> str:
    historicos = _historicos_ordenados(producao)
    finalizados = [getattr(h, "finalizado_em", None) for h in historicos if getattr(h, "finalizado_em", None)]
    if finalizados:
        return _formatar_hora(max(finalizados))
    return "N/A"


def _intercorrencia(agendamento: Agendamento, producao: Producao | None) -> str:
    if bool(getattr(agendamento, "tem_intercorrencia", False)):
        return "Sim"

    if producao and _texto_limpo(getattr(producao, "intercorrencias", None), ""):
        return "Sim"

    for historico in _historicos_ordenados(producao):
        if _texto_limpo(getattr(historico, "intercorrencia", None), ""):
            return "Sim"

    return "Não"


def _telefone_tutor(agendamento: Agendamento) -> str:
    cliente = getattr(agendamento, "cliente", None)
    telefone = getattr(cliente, "telefone", None) or getattr(cliente, "telefone_fixo", None)
    return _texto_limpo(telefone, "N/A")


def _montar_linha(agendamento: Agendamento, consumos_por_agendamento: set[int]) -> list[str]:
    cliente = getattr(agendamento, "cliente", None)
    pet = getattr(agendamento, "pet", None)
    producao = getattr(agendamento, "producao", None)
    etapas = _mapear_historicos_por_etapa(producao)

    responsavel_tosa = _nome_primeiro_responsavel(etapas, "TOSA")

    return [
        _formatar_data(getattr(agendamento, "data", None)),
        _formatar_hora(getattr(agendamento, "created_at", None)),
        _tipo_servico(agendamento.id, consumos_por_agendamento),
        _categoria_cliente(agendamento),
        _texto_limpo(getattr(cliente, "nome", None)),
        _telefone_tutor(agendamento),
        _texto_limpo(getattr(pet, "nome", None)),
        _texto_limpo(getattr(pet, "raca", None)),
        _texto_limpo(getattr(pet, "porte", None)),
        _ordem_servico(agendamento),
        _normalizar_prioridade(getattr(agendamento, "prioridade", None)),
        _inicio_linha(producao),
        _nome_primeiro_responsavel(etapas, "PRE_BANHO"),
        _nome_primeiro_responsavel(etapas, "BANHO"),
        _nome_primeiro_responsavel(etapas, "FINALIZACAO_BANHO"),
        _nome_primeiro_responsavel(etapas, "PRE_TOSA"),
        responsavel_tosa,
        responsavel_tosa,
        _horario_finalizacao(producao),
        _intercorrencia(agendamento, producao),
    ]


def _buscar_agendamentos(
    db: Session,
    empresa_id: int,
    data_inicial: date,
    data_final: date,
):
    return (
        db.query(Agendamento)
        .options(
            joinedload(Agendamento.cliente),
            joinedload(Agendamento.pet),
            joinedload(Agendamento.servicos_agendamento).joinedload(AgendamentoServico.servico),
            joinedload(Agendamento.producao)
            .joinedload(Producao.historicos)
            .joinedload(ProducaoHistorico.funcionario),
        )
        .filter(
            Agendamento.empresa_id == empresa_id,
            Agendamento.data >= data_inicial,
            Agendamento.data <= data_final,
        )
        .order_by(Agendamento.data.asc(), Agendamento.hora.asc(), Agendamento.id.asc())
        .all()
    )


def _buscar_consumos_assinatura_por_agendamento(
    db: Session,
    empresa_id: int,
    agendamento_ids: list[int],
) -> set[int]:
    if not agendamento_ids:
        return set()

    rows = (
        db.query(AssinaturaPetConsumo.agendamento_id)
        .filter(
            AssinaturaPetConsumo.empresa_id == empresa_id,
            AssinaturaPetConsumo.agendamento_id.in_(agendamento_ids),
            AssinaturaPetConsumo.status.in_(["PENDENTE", "CONFIRMADO"]),
        )
        .distinct()
        .all()
    )

    return {row[0] for row in rows if row[0]}


def _ajustar_larguras(ws):
    larguras = {
        "A": 13,
        "B": 18,
        "C": 20,
        "D": 20,
        "E": 32,
        "F": 18,
        "G": 20,
        "H": 28,
        "I": 18,
        "J": 34,
        "K": 14,
        "L": 18,
        "M": 24,
        "N": 24,
        "O": 24,
        "P": 24,
        "Q": 24,
        "R": 24,
        "S": 22,
        "T": 16,
    }

    for coluna, largura in larguras.items():
        ws.column_dimensions[coluna].width = largura


def _montar_planilha(linhas: list[list[str]]) -> BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "Banho e Tosa"

    ws.append(COLUNAS_RELATORIO)

    header_fill = PatternFill(fill_type="solid", fgColor="1F4E78")
    header_font = Font(bold=True, color="FFFFFF")
    center = Alignment(horizontal="center", vertical="center")
    left = Alignment(horizontal="left", vertical="center")

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center

    for linha in linhas:
        ws.append(linha)

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for idx, cell in enumerate(row, start=1):
            if idx in {1, 2, 11, 12, 19, 20}:
                cell.alignment = center
            else:
                cell.alignment = left

    ws.freeze_panes = "A2"
    _ajustar_larguras(ws)

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


@router.get("/exportar")
def exportar_relatorio_banho_tosa(
    data_inicial: str = Query(..., description="Data inicial no formato YYYY-MM-DD"),
    data_final: str = Query(..., description="Data final no formato YYYY-MM-DD"),
    db: Session = Depends(get_db),
    empresa_id: int = Depends(get_empresa_id_atual),
):
    dt_inicial = _parse_date_param(data_inicial, "data_inicial")
    dt_final = _parse_date_param(data_final, "data_final")

    if dt_final < dt_inicial:
        raise HTTPException(
            status_code=400,
            detail="data_final não pode ser menor que data_inicial.",
        )

    agendamentos = _buscar_agendamentos(
        db=db,
        empresa_id=empresa_id,
        data_inicial=dt_inicial,
        data_final=dt_final,
    )

    agendamento_ids = [ag.id for ag in agendamentos]
    consumos_por_agendamento = _buscar_consumos_assinatura_por_agendamento(
        db=db,
        empresa_id=empresa_id,
        agendamento_ids=agendamento_ids,
    )

    linhas = [_montar_linha(agendamento, consumos_por_agendamento) for agendamento in agendamentos]
    arquivo = _montar_planilha(linhas)

    nome_arquivo = (
        f"relatorio_banho_tosa_{dt_inicial.strftime('%Y%m%d')}_{dt_final.strftime('%Y%m%d')}.xlsx"
    )

    headers = {
        "Content-Disposition": f'attachment; filename="{nome_arquivo}"'
    }

    return StreamingResponse(
        arquivo,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers,
    )